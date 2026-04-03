#!/usr/bin/env python3
"""cuesheet-creator — turn a single video into a collaborative cue sheet."""
from __future__ import annotations

__version__ = "1.0.0"

import argparse
import datetime as dt
import importlib
import json
import math
import os
from pathlib import Path
import shutil
import subprocess
import sys
from typing import Any

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


REQUIRED_PACKAGES = {
    "opencv-python-headless": "cv2",
    "numpy": "numpy",
    "Pillow": "PIL",
    "openpyxl": "openpyxl",
}

OPTIONAL_COMPONENTS = {
    "scene:scenedetect": {
        "group": "scene",
        "pip_name": "scenedetect[opencv]",
        "pip_spec": "scenedetect[opencv]>=0.6,<1",
        "import_name": "scenedetect",
    },
    "asr:faster-whisper": {
        "group": "asr",
        "pip_name": "faster-whisper",
        "pip_spec": "faster-whisper>=1.0,<2",
        "import_name": "faster_whisper",
    },
    "ocr:rapidocr": {
        "group": "ocr",
        "pip_name": "rapidocr-onnxruntime",
        "pip_spec": "rapidocr-onnxruntime>=1.3,<2",
        "import_name": "rapidocr_onnxruntime",
    },
    "ocr:easyocr": {
        "group": "ocr-extra",
        "pip_name": "easyocr",
        "pip_spec": "easyocr>=1.7,<2",
        "import_name": "easyocr",
    },
    "ocr:paddleocr": {
        "group": "ocr-extra",
        "pip_name": "paddleocr",
        "pip_spec": "paddleocr>=2.7,<3",
        "import_name": "paddleocr",
    },
}

TEMPLATE_COLUMNS = {
    "script": [
        "shot_block",
        "start_time",
        "end_time",
        "scene",
        "location",
        "characters",
        "event",
        "important_dialogue",
        "confidence",
        "needs_confirmation",
    ],
    "production": [
        "shot_block",
        "start_time",
        "end_time",
        "keyframe",
        "shot_size",
        "angle_or_lens",
        "motion",
        "scene",
        "mood",
        "location",
        "characters",
        "event",
        "important_dialogue",
        "music_note",
        "director_note",
        "confidence",
        "needs_confirmation",
    ],
    "music-director": [
        "shot_block",
        "start_time",
        "end_time",
        "mood",
        "event",
        "important_dialogue",
        "music_note",
        "rhythm_change",
        "instrumentation",
        "dynamics",
        "confidence",
        "needs_confirmation",
    ],
}

DEFAULT_COLUMN_WIDTHS = {
    "shot_block": 12,
    "start_time": 14,
    "end_time": 14,
    "keyframe": 24,
    "shot_size": 12,
    "angle_or_lens": 20,
    "motion": 18,
    "scene": 18,
    "mood": 20,
    "location": 18,
    "characters": 20,
    "event": 28,
    "important_dialogue": 30,
    "music_note": 28,
    "director_note": 28,
    "confidence": 18,
    "needs_confirmation": 24,
    "rhythm_change": 20,
    "instrumentation": 20,
    "dynamics": 18,
}


SUPPORTED_OPTIONAL_GROUPS = {"asr", "ocr", "ocr-extra", "scene"}

PREPARE_ENV_MODES = {
    "check-only": set(),
    "install-required": set(),
    "install-scene": {"scene"},
    "install-asr": {"asr"},
    "install-ocr": {"ocr"},
    "install-all": set(SUPPORTED_OPTIONAL_GROUPS),
}

PREPARE_ENV_DEFAULT_FILES = {
    "precheck": "selfcheck.pre.json",
    "install_report": "install_report.json",
    "postcheck": "selfcheck.post.json",
    "report": "prepare_env.json",
}

SKILL_ROOT = Path(__file__).resolve().parent.parent
LOCAL_FFMPEG_BIN_ENV = "CUESHEET_CREATOR_FFMPEG_BIN_DIR"
LOCAL_FFMPEG_SEARCH_ROOT = SKILL_ROOT / "tools" / "ffmpeg"

# Runtime overrides set by --ffmpeg-path / --ffprobe-path CLI args.
# These are populated in main() before any command runs.
_CLI_COMMAND_OVERRIDES: dict[str, str] = {}


def ensure_parent(path: Path) -> None:



    path.parent.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, data: dict[str, Any]) -> None:
    ensure_parent(path)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def format_seconds(seconds: float) -> str:
    if seconds < 0:
        seconds = 0.0
    whole = int(seconds)
    millis = int(round((seconds - whole) * 1000))
    if millis == 1000:
        whole += 1
        millis = 0
    hours = whole // 3600
    minutes = (whole % 3600) // 60
    secs = whole % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}.{millis:03d}"


def seconds_from_timecode(text: str) -> float:
    """Parse a timecode string into seconds.

    Accepted formats:
      HH:MM:SS.mmm   (canonical)
      HH:MM:SS,mmm   (SRT-style comma separator)
      HH:MM:SS        (no milliseconds — treated as .000)
      MM:SS.mmm
      MM:SS
      SS.mmm
      SS

    Raises ValueError with a user-friendly message on unrecognized input.
    """
    raw = text.strip()
    if not raw:
        raise ValueError("Empty timecode string.")

    # Normalize comma separator (SRT-style) to dot
    raw = raw.replace(",", ".")

    parts = raw.split(":")
    try:
        if len(parts) == 3:
            hh = int(parts[0])
            mm = int(parts[1])
            sec_parts = parts[2].split(".")
            ss = int(sec_parts[0])
            ms = int(sec_parts[1].ljust(3, "0")[:3]) if len(sec_parts) > 1 else 0
            return hh * 3600 + mm * 60 + ss + ms / 1000.0
        if len(parts) == 2:
            mm = int(parts[0])
            sec_parts = parts[1].split(".")
            ss = int(sec_parts[0])
            ms = int(sec_parts[1].ljust(3, "0")[:3]) if len(sec_parts) > 1 else 0
            return mm * 60 + ss + ms / 1000.0
        if len(parts) == 1:
            sec_parts = parts[0].split(".")
            ss = int(sec_parts[0])
            ms = int(sec_parts[1].ljust(3, "0")[:3]) if len(sec_parts) > 1 else 0
            return ss + ms / 1000.0
    except (ValueError, IndexError):
        pass

    raise ValueError(
        f"Unrecognized timecode format: '{text}'. "
        f"Expected HH:MM:SS.mmm, HH:MM:SS, MM:SS.mmm, MM:SS, or bare seconds."
    )


def run_command(command: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        command, capture_output=True, text=True, check=False,
        encoding="utf-8", errors="replace",
    )


def truncate_text(text: str | None, limit: int = 4000) -> str:
    if not text:
        return ""
    text = text.strip()
    if len(text) <= limit:
        return text
    return text[:limit] + "\n...[truncated]"


def resolved_path(value: str) -> str:
    """Argparse type helper: expand ~ and resolve to absolute path."""
    return str(Path(value).expanduser().resolve())


def detect_platform_family() -> str:
    if sys.platform.startswith("win"):
        return "windows"
    if sys.platform == "darwin":
        return "macos"
    return "linux"


def command_filename(name: str) -> str:
    if detect_platform_family() == "windows" and not name.lower().endswith(".exe"):
        return f"{name}.exe"
    return name


def _common_ffmpeg_dirs() -> list[tuple[Path, str]]:
    """Return platform-specific common FFmpeg install locations."""
    family = detect_platform_family()
    dirs: list[tuple[Path, str]] = []
    if family == "windows":
        # winget typical location
        local_app = Path(os.environ.get("LOCALAPPDATA", "")) / "Microsoft" / "WinGet" / "Links"
        if local_app.exists():
            dirs.append((local_app, "winget"))
        # scoop
        scoop_root = Path(os.environ.get("SCOOP", Path.home() / "scoop"))
        dirs.append((scoop_root / "shims", "scoop"))
        dirs.append((scoop_root / "apps" / "ffmpeg" / "current" / "bin", "scoop"))
        # choco
        choco_root = Path(os.environ.get("ChocolateyInstall", r"C:\ProgramData\chocolatey"))
        dirs.append((choco_root / "bin", "choco"))
        # Common manual extract locations
        for drive in ("C:", "D:"):
            dirs.append((Path(drive) / "ffmpeg" / "bin", "manual"))
            dirs.append((Path(drive) / "tools" / "ffmpeg" / "bin", "manual"))
        prog = Path(os.environ.get("ProgramFiles", r"C:\Program Files"))
        dirs.append((prog / "ffmpeg" / "bin", "programfiles"))
    elif family == "macos":
        dirs.append((Path("/opt/homebrew/bin"), "homebrew-arm"))
        dirs.append((Path("/usr/local/bin"), "homebrew-intel"))
        dirs.append((Path("/opt/local/bin"), "macports"))
    else:
        dirs.append((Path("/usr/bin"), "system"))
        dirs.append((Path("/usr/local/bin"), "local"))
        dirs.append((Path("/snap/bin"), "snap"))
    return dirs


def iter_local_ffmpeg_bin_dirs() -> list[tuple[Path, str]]:
    """Ordered list of directories to search for ffmpeg/ffprobe.
    Priority: explicit override > skill-local tools/ > common OS locations > PATH."""
    candidates: list[tuple[Path, str]] = []
    seen: set[str] = set()

    def add(path: Path | None, source: str) -> None:
        if path is None:
            return
        key = str(path.resolve()) if path.exists() else str(path)
        if key in seen:
            return
        seen.add(key)
        candidates.append((path, source))

    # 1. Explicit env var override (highest priority)
    env_value = os.environ.get(LOCAL_FFMPEG_BIN_ENV, "").strip()
    if env_value:
        add(Path(env_value).expanduser(), f"env:{LOCAL_FFMPEG_BIN_ENV}")

    # 2. Skill-local tools/ffmpeg/ directory
    add(LOCAL_FFMPEG_SEARCH_ROOT / "bin", "local-tools")
    add(LOCAL_FFMPEG_SEARCH_ROOT, "local-tools")
    if LOCAL_FFMPEG_SEARCH_ROOT.exists():
        for child in sorted(LOCAL_FFMPEG_SEARCH_ROOT.iterdir(), key=lambda item: item.name.lower()):
            if child.is_dir():
                add(child / "bin", f"local-tools:{child.name}")
                add(child, f"local-tools:{child.name}")

    # 3. Common OS-level install locations (winget/scoop/choco/homebrew/etc.)
    for common_dir, source in _common_ffmpeg_dirs():
        add(common_dir, source)

    return candidates


def resolve_command_path(name: str) -> tuple[str | None, str | None]:
    """Resolve path for an external command.
    Priority: CLI override > env var > skill-local tools/ > common OS locations > PATH."""
    # Highest priority: explicit --ffmpeg-path / --ffprobe-path from CLI
    cli_override = _CLI_COMMAND_OVERRIDES.get(name)
    if cli_override:
        p = Path(cli_override)
        if p.exists() and p.is_file():
            return str(p.resolve()), "cli-override"
        # CLI override was given but path doesn't exist — print a clear warning
        print(f"WARNING: --{name}-path was set to '{cli_override}' but that file does not exist. Falling back to auto-detection.", file=sys.stderr)
        # Fall through to normal auto-detection instead of hard-failing

    executable = command_filename(name)
    for bin_dir, source in iter_local_ffmpeg_bin_dirs():
        candidate = bin_dir / executable
        if candidate.exists() and candidate.is_file():
            return str(candidate.resolve()), source
    path = shutil.which(name)
    if path:
        return path, "PATH"
    return None, None


def ffmpeg_install_hints() -> dict[str, Any]:
    """Return structured install hints: primary (give first) + fallbacks (give only on failure)."""
    family = detect_platform_family()
    if family == "windows":
        return {
            "primary": "Download FFmpeg from https://www.gyan.dev/ffmpeg/builds/ (essentials build zip). Extract to <skill-root>/tools/ffmpeg/ so it contains <release-folder>/bin/ffmpeg.exe. No PATH changes needed — cuesheet-creator auto-detects it.",
            "fallbacks": [
                "If portable placement didn't work: check that <skill-root>/tools/ffmpeg/ contains either bin/ffmpeg.exe directly or <release-folder>/bin/ffmpeg.exe one level down.",
                "Alternative: install via winget (`winget install Gyan.FFmpeg`), scoop, or choco. After install, close and reopen your PowerShell window, then re-check.",
                f"Direct override: pass `--ffmpeg-path <path> --ffprobe-path <path>` to any command, or set {LOCAL_FFMPEG_BIN_ENV}=<directory>.",
            ],
        }
    if family == "macos":
        return {
            "primary": "Run `brew install ffmpeg` in Terminal. cuesheet-creator auto-detects /opt/homebrew/bin and /usr/local/bin.",
            "fallbacks": [
                "If brew install succeeded but selfcheck still shows MISSING: close and reopen your Terminal window, then re-check.",
                "Alternative: extract FFmpeg binaries into <skill-root>/tools/ffmpeg/bin/.",
                "Direct override: pass `--ffmpeg-path` / `--ffprobe-path` to any command.",
            ],
        }
    return {
        "primary": "Install via package manager: `sudo apt install ffmpeg` (or your distro's equivalent).",
        "fallbacks": [
            "If install succeeded but selfcheck still shows MISSING: close and reopen your terminal, then re-check.",
            "Alternative: extract FFmpeg binaries into <skill-root>/tools/ffmpeg/bin/.",
            "Direct override: pass `--ffmpeg-path` / `--ffprobe-path` to any command.",
        ],
    }



def normalize_optional_groups(value: str | None) -> set[str]:
    if not value:
        return set()
    raw = value.strip().lower()
    if raw in {"", "none"}:
        return set()
    groups: set[str] = set()
    for part in raw.split(","):
        item = part.strip()
        if not item or item == "none":
            continue
        if item == "all":
            return set(SUPPORTED_OPTIONAL_GROUPS)
        if item not in SUPPORTED_OPTIONAL_GROUPS:
            valid = ", ".join(sorted(SUPPORTED_OPTIONAL_GROUPS | {"all", "none"}))
            raise ValueError(f"Unknown optional group: {item}, valid values: {valid}")
        groups.add(item)
    return groups


def unique_in_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            ordered.append(item)
    return ordered


def check_command(name: str) -> dict[str, Any]:
    path, source = resolve_command_path(name)
    result: dict[str, Any] = {
        "name": name,
        "available": bool(path),
        "path": path,
        "source": source,
        "version": None,
    }
    if not path:
        return result

    completed = run_command([path, "-version"])
    text = (completed.stdout or completed.stderr or "").strip().splitlines()
    result["version"] = text[0] if text else None
    return result



def check_module(package_name: str, import_name: str, pip_name: str | None = None, pip_spec: str | None = None, group: str | None = None) -> dict[str, Any]:
    try:
        mod = importlib.import_module(import_name)
        version = getattr(mod, "__version__", None)
        return {
            "package": package_name,
            "pip_name": pip_name or package_name,
            "pip_spec": pip_spec or pip_name or package_name,
            "group": group,
            "import": import_name,
            "available": True,
            "installed_version": version,
        }
    except Exception as exc:
        return {
            "package": package_name,
            "pip_name": pip_name or package_name,
            "pip_spec": pip_spec or pip_name or package_name,
            "group": group,
            "import": import_name,
            "available": False,
            "installed_version": None,
            "error": str(exc),
        }


def summarize_report(required_packages: dict[str, Any], optional_components: dict[str, Any], ffmpeg: dict[str, Any], ffprobe: dict[str, Any]) -> dict[str, Any]:
    missing_required_packages = [package for package, info in required_packages.items() if not info["available"]]
    missing_optional_components = [name for name, info in optional_components.items() if not info["available"]]
    missing_external_commands = []
    if not ffmpeg["available"]:
        missing_external_commands.append("ffmpeg")
    if not ffprobe["available"]:
        missing_external_commands.append("ffprobe")
    return {
        "missing_required_packages": missing_required_packages,
        "missing_optional_components": missing_optional_components,
        "missing_external_commands": missing_external_commands,
    }


def make_selfcheck_report() -> dict[str, Any]:
    required_packages = {
        package: check_module(package, import_name, pip_name=package)
        for package, import_name in REQUIRED_PACKAGES.items()
    }
    optional_components = {
        name: check_module(
            package_name=name,
            import_name=meta["import_name"],
            pip_name=meta["pip_name"],
            pip_spec=meta.get("pip_spec"),
            group=meta["group"],
        )
        for name, meta in OPTIONAL_COMPONENTS.items()
    }

    ffmpeg = check_command("ffmpeg")
    ffprobe = check_command("ffprobe")

    blocking: list[str] = []
    warnings: list[str] = []

    if sys.version_info < (3, 10):
        blocking.append("Python version below 3.10")

    if not ffmpeg["available"]:
        blocking.append("Missing ffmpeg")
    if not ffprobe["available"]:
        blocking.append("Missing ffprobe")

    for package, item in required_packages.items():
        if not item["available"]:
            blocking.append(f"Missing required Python package: {package}")

    for name, item in optional_components.items():
        if not item["available"] and item.get("group") != "ocr-extra":
            warnings.append(f"Optional component unavailable: {name}")

    summary = summarize_report(required_packages, optional_components, ffmpeg, ffprobe)

    return {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "python": {
            "available": True,
            "version": sys.version.split()[0],
            "executable": sys.executable,
            "ok": sys.version_info >= (3, 10),
        },
        "commands": {
            "ffmpeg": ffmpeg,
            "ffprobe": ffprobe,
        },
        "required_packages": required_packages,
        "optional_components": optional_components,
        "summary": summary,
        "install_options": {
            "check_only": "python scripts/cuesheet_creator.py selfcheck --json-out <out-dir>/selfcheck.json",
            "install_required": "python scripts/cuesheet_creator.py install-deps --include-optional none --report-out <out-dir>/install_report.json",
            "install_with_scene": "python scripts/cuesheet_creator.py install-deps --include-optional scene --report-out <out-dir>/install_report.json",
            "install_with_asr": "python scripts/cuesheet_creator.py install-deps --include-optional asr --report-out <out-dir>/install_report.json",
            "install_with_ocr": "python scripts/cuesheet_creator.py install-deps --include-optional ocr --report-out <out-dir>/install_report.json",
            "install_with_all_optional": "python scripts/cuesheet_creator.py install-deps --include-optional all --report-out <out-dir>/install_report.json",
            "prepare_check_only": "python scripts/cuesheet_creator.py prepare-env --mode check-only --out-dir <out-dir>",
            "prepare_install_required": "python scripts/cuesheet_creator.py prepare-env --mode install-required --out-dir <out-dir>",
            "prepare_install_scene": "python scripts/cuesheet_creator.py prepare-env --mode install-scene --out-dir <out-dir>",
            "prepare_install_asr": "python scripts/cuesheet_creator.py prepare-env --mode install-asr --out-dir <out-dir>",
            "prepare_install_ocr": "python scripts/cuesheet_creator.py prepare-env --mode install-ocr --out-dir <out-dir>",
            "prepare_install_all": "python scripts/cuesheet_creator.py prepare-env --mode install-all --out-dir <out-dir>",
        },

        "guidance": {
            "ffmpeg_install_hints": ffmpeg_install_hints(),
            "local_ffmpeg_search_root": str(LOCAL_FFMPEG_SEARCH_ROOT),
            "local_ffmpeg_bin_env": LOCAL_FFMPEG_BIN_ENV,
            "python_package_policy": [
                "Only run install-deps when the user explicitly allows auto-install.",
                "Prefer installing in isolated or managed Python environments.",
                "install-deps only handles Python packages; it does not auto-install ffmpeg/ffprobe.",
            ],
        },

        "overall": {
            "ready": len(blocking) == 0,
            "blocking": blocking,
            "warnings": warnings,
        },
    }


def print_selfcheck_text(report: dict[str, Any]) -> None:
    print("=== cuesheet-creator selfcheck ===")
    print(f"Python: {report['python']['version']} @ {report['python']['executable']}")
    ffmpeg_state = report["commands"]["ffmpeg"]
    ffprobe_state = report["commands"]["ffprobe"]
    print(f"ffmpeg : {'OK' if ffmpeg_state['available'] else 'MISSING'}" + (f" [{ffmpeg_state.get('source')}]" if ffmpeg_state.get('source') else ""))
    if ffmpeg_state.get("path"):
        print(f"         {ffmpeg_state['path']}")
    print(f"ffprobe: {'OK' if ffprobe_state['available'] else 'MISSING'}" + (f" [{ffprobe_state.get('source')}]" if ffprobe_state.get('source') else ""))
    if ffprobe_state.get("path"):
        print(f"         {ffprobe_state['path']}")
    print("Required packages:")

    for package, info in report["required_packages"].items():
        state = "OK" if info["available"] else "MISSING"
        version = info.get("installed_version") or "unknown"
        print(f"  - {package}: {state} ({version})")
    print("Optional components:")
    for name, info in report["optional_components"].items():
        state = "OK" if info["available"] else "UNAVAILABLE"
        version = info.get("installed_version") or "unknown"
        group = info.get("group") or "optional"
        print(f"  - {name} [{group}]: {state} ({version})")
    if report["overall"]["blocking"]:
        print("Blocking issues:")
        for item in report["overall"]["blocking"]:
            print(f"  - {item}")
    if report["overall"]["warnings"]:
        print("Warnings:")
        for item in report["overall"]["warnings"]:
            print(f"  - {item}")
    if report["summary"]["missing_external_commands"]:
        print("ffmpeg guidance:")
        print(f"  - local search root: {report['guidance']['local_ffmpeg_search_root']}")
        print(f"  - override env    : {report['guidance']['local_ffmpeg_bin_env']}")
        hints = report["guidance"]["ffmpeg_install_hints"]
        print(f"  - (primary) {hints['primary']}")
        for fb in hints.get("fallbacks", []):
            print(f"  - (fallback) {fb}")

    print("Install flow:")
    print(f"  - only selfcheck : {report['install_options']['check_only']}")
    print(f"  - auto required  : {report['install_options']['install_required']}")
    print(f"  - + scene detect : {report['install_options']['install_with_scene']}")
    print(f"  - + ASR optional : {report['install_options']['install_with_asr']}")
    print(f"  - + OCR optional : {report['install_options']['install_with_ocr']}")
    print(f"  - + all optional : {report['install_options']['install_with_all_optional']}")
    print("One-command flow:")
    print(f"  - check only     : {report['install_options']['prepare_check_only']}")
    print(f"  - install needed : {report['install_options']['prepare_install_required']}")
    print(f"  - + scene detect : {report['install_options']['prepare_install_scene']}")
    print(f"  - + ASR optional : {report['install_options']['prepare_install_asr']}")
    print(f"  - + OCR optional : {report['install_options']['prepare_install_ocr']}")
    print(f"  - install all    : {report['install_options']['prepare_install_all']}")

    print(f"Ready: {'YES' if report['overall']['ready'] else 'NO'}")



def cmd_selfcheck(args: argparse.Namespace) -> int:
    report = make_selfcheck_report()
    if args.json_out:
        write_json(Path(args.json_out), report)
    if args.output_format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print_selfcheck_text(report)
    if args.fail_on_missing_required and not report["overall"]["ready"]:
        return 1
    return 0


def collect_missing_python_packages(report: dict[str, Any], optional_groups: set[str]) -> list[str]:
    packages = [
        info.get("pip_spec") or info["pip_name"]
        for info in report["required_packages"].values()
        if not info["available"]
    ]
    for info in report["optional_components"].values():
        if info["available"]:
            continue
        if info.get("group") in optional_groups:
            packages.append(info.get("pip_spec") or info["pip_name"])
    return unique_in_order(packages)


def ensure_pip_available(python_executable: str) -> tuple[bool, list[dict[str, Any]]]:
    steps: list[dict[str, Any]] = []

    check_cmd = [python_executable, "-m", "pip", "--version"]
    check_result = run_command(check_cmd)
    steps.append(
        {
            "step": "check-pip",
            "command": check_cmd,
            "returncode": check_result.returncode,
            "stdout": truncate_text(check_result.stdout),
            "stderr": truncate_text(check_result.stderr),
        }
    )
    if check_result.returncode == 0:
        return True, steps

    ensure_cmd = [python_executable, "-m", "ensurepip", "--upgrade"]
    ensure_result = run_command(ensure_cmd)
    steps.append(
        {
            "step": "ensurepip",
            "command": ensure_cmd,
            "returncode": ensure_result.returncode,
            "stdout": truncate_text(ensure_result.stdout),
            "stderr": truncate_text(ensure_result.stderr),
        }
    )

    recheck_result = run_command(check_cmd)
    steps.append(
        {
            "step": "recheck-pip",
            "command": check_cmd,
            "returncode": recheck_result.returncode,
            "stdout": truncate_text(recheck_result.stdout),
            "stderr": truncate_text(recheck_result.stderr),
        }
    )
    return recheck_result.returncode == 0, steps


def load_requirements_constraints() -> dict[str, str]:
    """Parse requirements.txt and return {bare_package_name: 'name>=x,<y'} for uncommented lines."""
    req_path = Path(__file__).resolve().parent.parent / "requirements.txt"
    constraints: dict[str, str] = {}
    if not req_path.exists():
        return constraints
    for line in req_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        # Extract bare name (before any version specifier)
        bare = line.split(">=")[0].split("<=")[0].split("==")[0].split("<")[0].split(">")[0].split("[")[0].strip()
        if bare:
            constraints[bare.lower()] = line
    return constraints


def build_pip_install_command(args: argparse.Namespace, packages: list[str]) -> list[str]:
    constraints = load_requirements_constraints()
    resolved: list[str] = []
    for pkg in packages:
        # Look up version-constrained spec from requirements.txt
        bare = pkg.split(">=")[0].split("[")[0].strip().lower()
        if bare in constraints:
            resolved.append(constraints[bare])
        else:
            resolved.append(pkg)
    command = [sys.executable, "-m", "pip", "install"]
    if args.upgrade_pip:
        command.extend(["--upgrade", "pip"])
    command.extend(resolved)
    if args.index_url:
        command.extend(["--index-url", args.index_url])
    if args.extra_index_url:
        command.extend(["--extra-index-url", args.extra_index_url])
    return command


def print_install_report(report: dict[str, Any]) -> None:
    print("=== cuesheet-creator install-deps ===")
    print(f"Python: {report['python_executable']}")
    print(f"Optional groups: {', '.join(report['optional_groups']) if report['optional_groups'] else 'none'}")
    print(f"Dry run: {'YES' if report['dry_run'] else 'NO'}")
    print(f"Packages selected: {', '.join(report['packages_to_install']) if report['packages_to_install'] else '(none)'}")
    if report["precheck"]["summary"]["missing_external_commands"]:
        print("External blockers still require manual install:")
        for item in report["precheck"]["summary"]["missing_external_commands"]:
            print(f"  - {item}")
        hints = report["precheck"]["guidance"]["ffmpeg_install_hints"]
        print(f"  - (primary) {hints['primary']}")
        for fb in hints.get("fallbacks", []):
            print(f"  - (fallback) {fb}")
    if report.get("pip_bootstrap_steps"):
        print("pip bootstrap steps:")
        for step in report["pip_bootstrap_steps"]:
            print(f"  - {step['step']}: rc={step['returncode']}")
    if report.get("pip_command"):
        print("Install command:")
        print("  " + " ".join(report["pip_command"]))
    if report.get("pip_returncode") is not None:
        print(f"pip returncode: {report['pip_returncode']}")
    if report.get("pip_stdout"):
        print("pip stdout:")
        print(report["pip_stdout"])
    if report.get("pip_stderr"):
        print("pip stderr:")
        print(report["pip_stderr"])
    print(f"Ready after install: {'YES' if report['postcheck']['overall']['ready'] else 'NO'}")
    if report["postcheck"]["overall"]["blocking"]:
        print("Remaining blocking issues:")
        for item in report["postcheck"]["overall"]["blocking"]:
            print(f"  - {item}")


def run_install_deps_flow(args: argparse.Namespace) -> tuple[int, dict[str, Any]]:
    optional_groups = normalize_optional_groups(args.include_optional)
    precheck = make_selfcheck_report()
    packages_to_install = collect_missing_python_packages(precheck, optional_groups)

    report: dict[str, Any] = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "python_executable": sys.executable,
        "optional_groups": sorted(optional_groups),
        "dry_run": bool(args.dry_run),
        "precheck": precheck,
        "packages_to_install": packages_to_install,
        "pip_bootstrap_steps": [],
        "pip_command": None,
        "pip_returncode": None,
        "pip_stdout": "",
        "pip_stderr": "",
        "postcheck": precheck,
    }

    if not packages_to_install:
        return 0, report

    pip_ready, pip_bootstrap_steps = ensure_pip_available(sys.executable)
    report["pip_bootstrap_steps"] = pip_bootstrap_steps
    if not pip_ready:
        report["pip_stderr"] = "Unable to initialize pip. Please check the current Python environment."
        return 1, report

    pip_command = build_pip_install_command(args, packages_to_install)
    report["pip_command"] = pip_command

    if not args.dry_run:
        install_result = run_command(pip_command)
        report["pip_returncode"] = install_result.returncode
        report["pip_stdout"] = truncate_text(install_result.stdout, limit=12000)
        report["pip_stderr"] = truncate_text(install_result.stderr, limit=12000)
        report["postcheck"] = make_selfcheck_report()
    else:
        report["postcheck"] = precheck

    if args.dry_run:
        return 0, report
    if report["pip_returncode"] not in (0, None):
        return 1, report
    return 0, report


def cmd_install_deps(args: argparse.Namespace) -> int:
    exit_code, report = run_install_deps_flow(args)
    if args.report_out:
        write_json(Path(args.report_out), report)
    if args.output_format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print_install_report(report)
    if exit_code != 0:
        return exit_code
    if args.fail_on_blocking and not report["postcheck"]["overall"]["ready"]:
        return 1
    return 0


def resolve_prepare_mode(mode: str) -> set[str]:
    if mode not in PREPARE_ENV_MODES:
        valid = ", ".join(sorted(PREPARE_ENV_MODES))
        raise ValueError(f"Unknown prepare mode: {mode}, valid values: {valid}")
    return set(PREPARE_ENV_MODES[mode])


def resolve_prepare_env_output_paths(args: argparse.Namespace) -> dict[str, Path | None]:
    out_dir = Path(args.out_dir).resolve() if args.out_dir else None

    def pick(explicit_value: str | None, default_name: str) -> Path | None:
        if explicit_value:
            return Path(explicit_value)
        if out_dir is not None:
            return out_dir / default_name
        return None

    return {
        "out_dir": out_dir,
        "precheck": pick(args.selfcheck_out, PREPARE_ENV_DEFAULT_FILES["precheck"]),
        "install_report": pick(args.install_report_out, PREPARE_ENV_DEFAULT_FILES["install_report"]),
        "postcheck": pick(args.postcheck_out, PREPARE_ENV_DEFAULT_FILES["postcheck"]),
        "report": pick(args.report_out, PREPARE_ENV_DEFAULT_FILES["report"]),
    }


def stringify_output_paths(paths: dict[str, Path | None]) -> dict[str, str | None]:
    return {
        key: (str(value) if isinstance(value, Path) else None)
        for key, value in paths.items()
        if key != "out_dir"
    }


def print_prepare_env_report(report: dict[str, Any]) -> None:

    print("=== cuesheet-creator prepare-env ===")
    print(f"Mode: {report['mode']}")
    print(f"Dry run: {'YES' if report['dry_run'] else 'NO'}")
    if report.get("output_directory"):
        print(f"Output directory: {report['output_directory']}")
    if report.get("output_files"):
        print("Output files:")
        for key, value in report["output_files"].items():
            if value:
                print(f"  - {key}: {value}")
    print(f"Precheck ready : {'YES' if report['precheck']['overall']['ready'] else 'NO'}")

    if report["precheck"]["overall"]["blocking"]:
        print("Precheck blocking issues:")
        for item in report["precheck"]["overall"]["blocking"]:
            print(f"  - {item}")
    if report["install_invoked"]:
        print("Install step: EXECUTED")
        install_report = report.get("install_report") or {}
        packages = install_report.get("packages_to_install") or []
        print(f"  Packages selected: {', '.join(packages) if packages else '(none)'}")
        if install_report.get("pip_command"):
            print("  Install command:")
            print("    " + " ".join(install_report["pip_command"]))
        if install_report.get("pip_returncode") is not None:
            print(f"  pip returncode: {install_report['pip_returncode']}")
    else:
        print("Install step: SKIPPED")
    print(f"Postcheck ready: {'YES' if report['postcheck']['overall']['ready'] else 'NO'}")
    if report["postcheck"]["overall"]["blocking"]:
        print("Postcheck blocking issues:")
        for item in report["postcheck"]["overall"]["blocking"]:
            print(f"  - {item}")
    if report["postcheck"]["summary"]["missing_external_commands"]:
        print("External commands still need manual install:")
        for item in report["postcheck"]["summary"]["missing_external_commands"]:
            print(f"  - {item}")
        print(f"  - local search root: {report['postcheck']['guidance']['local_ffmpeg_search_root']}")
        print(f"  - override env    : {report['postcheck']['guidance']['local_ffmpeg_bin_env']}")
        hints = report["postcheck"]["guidance"]["ffmpeg_install_hints"]
        print(f"  - (primary) {hints['primary']}")
        for fb in hints.get("fallbacks", []):
            print(f"  - (fallback) {fb}")



def cmd_prepare_env(args: argparse.Namespace) -> int:
    optional_groups = resolve_prepare_mode(args.mode)
    output_paths = resolve_prepare_env_output_paths(args)
    precheck = make_selfcheck_report()
    if output_paths["precheck"]:
        write_json(output_paths["precheck"], precheck)

    output_files = stringify_output_paths(output_paths)
    if args.mode == "check-only":
        output_files["install_report"] = None

    report: dict[str, Any] = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "mode": args.mode,
        "optional_groups": sorted(optional_groups),
        "dry_run": bool(args.dry_run),
        "output_directory": str(output_paths["out_dir"]) if output_paths["out_dir"] else None,
        "output_files": output_files,
        "precheck": precheck,
        "install_invoked": False,
        "install_report": None,
        "postcheck": precheck,
    }


    exit_code = 0
    if args.mode != "check-only":
        install_args = argparse.Namespace(
            include_optional="none" if not optional_groups else ",".join(sorted(optional_groups)),
            dry_run=args.dry_run,
            report_out=None,
            output_format="json",
            index_url=args.index_url,
            extra_index_url=args.extra_index_url,
            upgrade_pip=args.upgrade_pip,
            fail_on_blocking=False,
        )
        install_exit_code, install_report = run_install_deps_flow(install_args)
        report["install_invoked"] = True
        report["install_report"] = install_report
        exit_code = install_exit_code
        if output_paths["install_report"]:
            write_json(output_paths["install_report"], install_report)
        if not args.dry_run:
            report["postcheck"] = make_selfcheck_report()
        else:
            report["postcheck"] = precheck
    else:
        report["postcheck"] = precheck

    if output_paths["postcheck"]:
        write_json(output_paths["postcheck"], report["postcheck"])
    if output_paths["report"]:
        write_json(output_paths["report"], report)

    if args.output_format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print_prepare_env_report(report)

    if exit_code != 0:
        return exit_code
    if args.fail_on_blocking and not report["postcheck"]["overall"]["ready"]:
        return 1
    return 0



def require_runtime_for_scan() -> tuple[Any, Any, Any]:

    ffprobe_path, _ffprobe_source = resolve_command_path("ffprobe")
    ffmpeg_path, _ffmpeg_source = resolve_command_path("ffmpeg")
    missing = []
    if not ffprobe_path:
        missing.append("ffprobe")
    if not ffmpeg_path:
        missing.append("ffmpeg")
    if missing:
        raise RuntimeError("Missing command: " + ", ".join(missing))

    missing_modules = []

    modules: dict[str, Any] = {}
    for import_name in ("cv2", "numpy", "PIL"):
        try:
            modules[import_name] = importlib.import_module(import_name)
        except Exception:
            missing_modules.append(import_name)
    if missing_modules:
        raise RuntimeError("Missing Python module: " + ", ".join(missing_modules))
    return modules["cv2"], modules["numpy"], modules["PIL"]


def ffprobe_metadata(video_path: Path) -> dict[str, Any]:
    ffprobe_path, _source = resolve_command_path("ffprobe")
    if not ffprobe_path:
        raise RuntimeError("Missing command: ffprobe")
    command = [
        ffprobe_path,
        "-v",
        "error",
        "-print_format",
        "json",
        "-show_format",
        "-show_streams",
        str(video_path),
    ]

    completed = run_command(command)
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or "ffprobe read failed")
    try:
        return json.loads(completed.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"ffprobe output not parseable: {exc}") from exc


def parse_fps(text: str | None) -> float | None:
    if not text or text in {"0/0", "N/A"}:
        return None
    if "/" in text:
        left, right = text.split("/", 1)
        denominator = float(right)
        if denominator == 0:
            return None
        return float(left) / denominator
    return float(text)


def safe_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value in (None, "", "N/A"):
            return default
        return float(value)
    except Exception:
        return default


def build_video_info(metadata: dict[str, Any]) -> dict[str, Any]:
    streams = metadata.get("streams", [])
    format_info = metadata.get("format", {})
    video_stream = next((s for s in streams if s.get("codec_type") == "video"), {})
    audio_streams = [s for s in streams if s.get("codec_type") == "audio"]

    width = int(video_stream.get("width") or 0)
    height = int(video_stream.get("height") or 0)
    fps = parse_fps(video_stream.get("avg_frame_rate")) or parse_fps(video_stream.get("r_frame_rate"))
    duration = safe_float(video_stream.get("duration"))
    if duration is None:
        duration = safe_float(format_info.get("duration"), 0.0) or 0.0

    return {
        "source_path": format_info.get("filename"),
        "format_name": format_info.get("format_name"),
        "duration_seconds": round(duration, 3),
        "duration_timecode": format_seconds(duration),
        "resolution": {
            "width": width,
            "height": height,
        },
        "fps": round(fps, 3) if fps else None,
        "video_codec": video_stream.get("codec_name"),
        "audio_tracks": len(audio_streams),
        "audio_codecs": [s.get("codec_name") for s in audio_streams if s.get("codec_name")],
        "bit_rate": safe_float(format_info.get("bit_rate")),
        "size_bytes": safe_float(format_info.get("size")),
    }


def compute_hist_distance(cv2: Any, np: Any, frame_a: Any, frame_b: Any) -> float:
    hist_a = []
    hist_b = []
    for channel in range(3):
        h1 = cv2.calcHist([frame_a], [channel], None, [32], [0, 256])
        h2 = cv2.calcHist([frame_b], [channel], None, [32], [0, 256])
        cv2.normalize(h1, h1)
        cv2.normalize(h2, h2)
        hist_a.append(h1)
        hist_b.append(h2)
    distances = []
    for h1, h2 in zip(hist_a, hist_b):
        corr = cv2.compareHist(h1, h2, cv2.HISTCMP_CORREL)
        distances.append(1.0 - max(min(corr, 1.0), -1.0))
    score = float(np.mean(distances))
    return max(score, 0.0)


def resize_frame(cv2: Any, frame: Any, max_width: int) -> Any:
    height, width = frame.shape[:2]
    if width <= max_width:
        return frame
    scale = max_width / float(width)
    new_size = (int(width * scale), int(height * scale))
    return cv2.resize(frame, new_size)


def read_frame_at(cv2: Any, capture: Any, seconds: float) -> Any | None:
    capture.set(cv2.CAP_PROP_POS_MSEC, max(seconds, 0.0) * 1000.0)
    success, frame = capture.read()
    if not success:
        return None
    return frame


def make_block_id(index: int) -> str:
    return f"A{index}"


def build_draft_blocks(scene_candidates: list[dict[str, Any]], duration: float) -> list[dict[str, Any]]:
    ordered = sorted(scene_candidates, key=lambda item: item["seconds"])
    deduped: list[dict[str, Any]] = []
    last_seconds: float | None = None
    for item in ordered:
        seconds = float(item["seconds"])
        if last_seconds is not None and abs(seconds - last_seconds) < 0.001:
            continue
        deduped.append(item)
        last_seconds = seconds
    if not deduped:
        return []

    blocks = []
    for idx, item in enumerate(deduped, start=1):
        start = float(item["seconds"])
        end = duration if idx == len(deduped) else float(deduped[idx]["seconds"])
        if end < start:
            end = start
        blocks.append(
            {
                "shot_block": make_block_id(idx),
                "start_seconds": round(start, 3),
                "start_time": format_seconds(start),
                "end_seconds": round(end, 3),
                "end_time": format_seconds(end),
                "keyframe": item.get("image_path"),
                "candidate_score": item.get("score"),
                "cut_reason": item.get("reason"),
                "visual_features": item.get("visual_features"),
            }
        )
    return blocks


def detect_scenes_scenedetect(video_path: Path, threshold: float) -> tuple[list[dict[str, Any]] | None, str | None]:
    """Use PySceneDetect ContentDetector if available. Returns (candidates, error_message).
    Returns (None, reason) if not installed or runtime fails."""
    try:
        from scenedetect import open_video, SceneManager
        from scenedetect.detectors import ContentDetector
    except ImportError:
        return None, "scenedetect not installed"

    try:
        video = open_video(str(video_path))
        manager = SceneManager()
        manager.add_detector(ContentDetector(threshold=threshold))
        manager.detect_scenes(video)
        scene_list = manager.get_scene_list()

        candidates: list[dict[str, Any]] = []
        candidates.append({
            "index": 1,
            "seconds": 0.0,
            "timecode": format_seconds(0.0),
            "score": 1.0,
            "reason": "start",
        })

        for idx, (start_tc, _end_tc) in enumerate(scene_list, start=2):
            seconds = start_tc.get_seconds()
            if seconds <= 0.001:
                continue
            candidates.append({
                "index": idx,
                "seconds": round(seconds, 3),
                "timecode": format_seconds(seconds),
                "score": 1.0,
                "reason": f"scenedetect_content>={threshold}",
            })

        return candidates, None
    except Exception as exc:
        return None, f"SceneDetect runtime error: {exc}"


def extract_audio_track(video_path: Path, out_dir: Path, start: float | None = None, end: float | None = None) -> tuple[Path | None, str | None]:
    """Extract first audio track to WAV using ffmpeg. Optionally clip to start/end seconds.
    Returns (path, error_message)."""
    ffmpeg_path, _source = resolve_command_path("ffmpeg")
    if not ffmpeg_path:
        return None, "Missing command: ffmpeg"
    audio_path = out_dir / "audio.wav"
    cmd = [ffmpeg_path, "-y"]

    if start is not None and start > 0.001:
        cmd.extend(["-ss", format_seconds(start)])
    cmd.extend(["-i", str(video_path)])
    if end is not None:
        duration = end - (start or 0.0)
        if duration > 0:
            cmd.extend(["-t", f"{duration:.3f}"])
    cmd.extend(["-vn", "-acodec", "pcm_s16le", "-ar", "16000", "-ac", "1", str(audio_path)])
    result = run_command(cmd)
    if result.returncode != 0:
        stderr_summary = truncate_text(result.stderr, limit=500)
        return None, f"ffmpeg audio extraction failed (rc={result.returncode}): {stderr_summary}"
    if not audio_path.exists():
        return None, "ffmpeg completed but audio.wav was not created"
    return audio_path, None


def run_asr_faster_whisper(audio_path: Path, model_size: str = "base") -> tuple[list[dict[str, Any]] | None, str | None]:
    """Run faster-whisper ASR. Returns (segments, error_message). Segments is None on failure."""
    try:
        from faster_whisper import WhisperModel
    except ImportError:
        return None, "faster-whisper not installed"

    try:
        model = WhisperModel(model_size, device="cpu", compute_type="int8")
        segments_iter, _info = model.transcribe(str(audio_path), beam_size=5)
        results: list[dict[str, Any]] = []
        for segment in segments_iter:
            results.append({
                "start": round(segment.start, 3),
                "end": round(segment.end, 3),
                "start_time": format_seconds(segment.start),
                "end_time": format_seconds(segment.end),
                "text": segment.text.strip(),
            })
        return results, None
    except Exception as exc:
        return None, f"ASR runtime error: {exc}"


def run_ocr_on_frames(frame_paths: list[str]) -> tuple[list[dict[str, Any]] | None, str | None]:
    """Run OCR on selected keyframes. Tries rapidocr > easyocr > paddleocr.
    Returns (detections, error_message). detections is None on failure."""
    ocr_engine = None
    engine_name = None
    init_notes: list[str] = []

    try:
        from rapidocr_onnxruntime import RapidOCR
        ocr_engine = RapidOCR()
        engine_name = "rapidocr"
    except Exception as exc:
        init_notes.append(f"rapidocr init failed: {exc}")

    if ocr_engine is None:
        try:
            import easyocr
            reader = easyocr.Reader(["ch_sim", "en"], gpu=False)
            ocr_engine = reader
            engine_name = "easyocr"
        except Exception as exc:
            init_notes.append(f"easyocr init failed: {exc}")

    if ocr_engine is None:
        try:
            from paddleocr import PaddleOCR
            ocr_engine = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)
            engine_name = "paddleocr"
        except Exception as exc:
            init_notes.append(f"paddleocr init failed: {exc}")

    if ocr_engine is None:
        detail = "; ".join(init_notes) if init_notes else "no engines installed"
        return None, f"No OCR engine available ({detail})"

    results: list[dict[str, Any]] = []
    frame_errors: list[str] = []
    for frame_path in frame_paths:
        texts: list[str] = []
        try:
            if engine_name == "rapidocr":
                result, _elapse = ocr_engine(frame_path)
                if result:
                    texts = [item[1] for item in result if item[1]]
            elif engine_name == "easyocr":
                raw = ocr_engine.readtext(frame_path)
                texts = [item[1] for item in raw if item[1]]
            elif engine_name == "paddleocr":
                result = ocr_engine.ocr(frame_path, cls=True)
                if result and result[0]:
                    texts = [line[1][0] for line in result[0] if line[1] and line[1][0]]
        except Exception as exc:
            frame_errors.append(f"{Path(frame_path).name}: {exc}")
            continue

        if texts:
            results.append({
                "frame": frame_path,
                "texts": texts,
                "engine": engine_name,
            })
    if frame_errors:
        error_summary = f"OCR failed on {len(frame_errors)} frame(s): {'; '.join(frame_errors[:5])}"
        if len(frame_errors) > 5:
            error_summary += f" ... and {len(frame_errors) - 5} more"
        return (results, error_summary)
    return (results, None) if results else ([], None)


def compute_frame_sharpness(cv2: Any, frame: Any) -> float:
    """Compute Laplacian variance as a sharpness score."""
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    return float(cv2.Laplacian(gray, cv2.CV_64F).var())


def compute_visual_features(cv2: Any, np: Any, frame: Any) -> dict[str, Any]:
    """Compute objective visual features from a single frame.
    Returns brightness, contrast, saturation, dominant hue, and derived labels."""
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    brightness = float(np.mean(gray))
    contrast = float(np.std(gray))
    saturation = float(np.mean(hsv[:, :, 1]))
    dominant_hue = float(np.mean(hsv[:, :, 0]))
    # Derive human-readable labels
    if brightness < 80:
        tone = "dark"
    elif brightness > 170:
        tone = "bright"
    else:
        tone = "mid"
    if 10 < dominant_hue < 30:
        color_temp = "warm"
    elif 90 < dominant_hue < 130:
        color_temp = "cool"
    else:
        color_temp = "neutral"
    return {
        "brightness": round(brightness, 1),
        "contrast": round(contrast, 1),
        "saturation": round(saturation, 1),
        "dominant_hue": round(dominant_hue, 1),
        "tone": tone,
        "color_temp": color_temp,
    }


def score_keyframe_candidates(
    cv2: Any,
    frames: list[tuple[float, Any, str]],
) -> list[dict[str, Any]]:
    """Score candidate frames by sharpness. Returns sorted list (best first)."""
    scored: list[dict[str, Any]] = []
    for seconds, frame, image_path in frames:
        sharpness = compute_frame_sharpness(cv2, frame)
        scored.append({
            "seconds": seconds,
            "image_path": image_path,
            "sharpness": round(sharpness, 2),
        })
    scored.sort(key=lambda x: x["sharpness"], reverse=True)
    return scored


def cmd_scan_video(args: argparse.Namespace) -> int:
    cv2, np, _pil = require_runtime_for_scan()

    video_path = Path(args.video)
    if not video_path.exists() or not video_path.is_file():
        raise FileNotFoundError(f"Video not found: {video_path}")

    # Default out-dir: same folder as the video, named <video-stem>_cuesheet/
    if args.out_dir:
        out_dir = Path(args.out_dir)
    else:
        out_dir = video_path.parent / f"{video_path.stem}_cuesheet"

    keyframe_dir = out_dir / "keyframes"
    keyframe_dir.mkdir(parents=True, exist_ok=True)

    probe = ffprobe_metadata(video_path)
    video_info = build_video_info(probe)

    capture = cv2.VideoCapture(str(video_path))
    if not capture.isOpened():
        raise RuntimeError("OpenCV cannot open video")

    fps = video_info.get("fps") or capture.get(cv2.CAP_PROP_FPS) or 0.0
    frame_count = int(capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    duration = video_info.get("duration_seconds") or 0.0
    if (not duration or duration <= 0) and fps and frame_count:
        duration = frame_count / float(fps)
        video_info["duration_seconds"] = round(duration, 3)
        video_info["duration_timecode"] = format_seconds(duration)
    video_info["frame_count"] = frame_count

    if duration <= 0:
        raise RuntimeError("Cannot determine video duration")

    # --- Clip range support ---
    effective_start = 0.0
    effective_end = duration
    if args.start_time:
        effective_start = seconds_from_timecode(args.start_time)
    if args.end_time:
        effective_end = seconds_from_timecode(args.end_time)
    effective_start = max(0.0, min(effective_start, duration))
    effective_end = max(effective_start, min(effective_end, duration))
    effective_duration = effective_end - effective_start
    if effective_duration <= 0:
        raise RuntimeError(f"Effective clip range is zero: {format_seconds(effective_start)} - {format_seconds(effective_end)}")

    sample_interval = max(float(args.sample_interval), 0.2)
    max_samples = int(args.max_samples) if args.max_samples else None

    sampled_frames: list[dict[str, Any]] = []
    scene_candidates: list[dict[str, Any]] = []
    notes: list[str] = []
    detection_method = "histogram"

    # --- Phase 2.1: Try scenedetect first, fall back to histogram ---
    sd_threshold = float(args.scene_threshold)
    sd_content_threshold = float(args.content_threshold) if hasattr(args, "content_threshold") and args.content_threshold else 27.0

    sd_candidates, sd_err = detect_scenes_scenedetect(video_path, sd_content_threshold)
    if sd_candidates is not None:
        detection_method = "scenedetect"
        # Filter to effective clip range and renumber
        filtered = [
            c for c in sd_candidates
            if c["seconds"] >= effective_start - 0.05 and c["seconds"] <= effective_end + 0.05
        ]
        if not filtered or filtered[0]["seconds"] > effective_start + 0.1:
            filtered.insert(0, {
                "index": 1,
                "seconds": round(effective_start, 3),
                "timecode": format_seconds(effective_start),
                "score": 1.0,
                "reason": "start",
            })
        for i, c in enumerate(filtered, start=1):
            c["index"] = i
        scene_candidates = filtered
        notes.append(f"Using PySceneDetect ContentDetector (threshold={sd_content_threshold})")
    else:
        notes.append(f"PySceneDetect unavailable ({sd_err}), falling back to histogram-based cut detection")
        notes.append("TIP: For better scene detection (especially dissolves/fades), install scenedetect: python scripts/cuesheet_creator.py prepare-env --mode install-scene --out-dir <out-dir>")

    # --- Sample frames and save keyframes ---
    prev_frame = None
    times: list[float] = []
    current = effective_start
    while current < effective_end:
        times.append(round(current, 3))
        current += sample_interval
    if not times or abs(times[-1] - effective_end) > 0.05:
        times.append(round(max(effective_end - 0.001, 0.0), 3))

    if max_samples and len(times) > max_samples:
        stride = math.ceil(len(times) / max_samples)
        times = times[::stride]
        if times[-1] < effective_end - 0.05:
            times.append(round(max(effective_end - 0.001, 0.0), 3))
        notes.append(f"Too many sample points, downsampled with stride={stride}")

    for idx, seconds in enumerate(times, start=1):
        frame = read_frame_at(cv2, capture, seconds)
        if frame is None:
            notes.append(f"{format_seconds(seconds)} frame extraction failed, skipped")
            continue
        frame = resize_frame(cv2, frame, max_width=args.max_width)
        image_name = f"frame_{idx:04d}_{int(seconds * 1000):010d}.jpg"
        image_path = keyframe_dir / image_name
        ok = cv2.imwrite(str(image_path), frame)
        if not ok:
            raise RuntimeError(f"Failed to write keyframe: {image_path}")

        sharpness = round(compute_frame_sharpness(cv2, frame), 2)
        visual_features = compute_visual_features(cv2, np, frame)
        score = 1.0 if prev_frame is None else round(compute_hist_distance(cv2, np, prev_frame, frame), 4)
        frame_record = {
            "index": idx,
            "seconds": round(seconds, 3),
            "timecode": format_seconds(seconds),
            "image_path": str(image_path),
            "score_from_previous": score,
            "sharpness": sharpness,
            "visual_features": visual_features,
        }
        sampled_frames.append(frame_record)

        # If using histogram fallback, build scene_candidates here
        if detection_method == "histogram":
            if prev_frame is None or score >= sd_threshold:
                reason = "start" if prev_frame is None else f"hist_diff>={sd_threshold}"
                scene_candidates.append(
                    {
                        "index": len(scene_candidates) + 1,
                        "seconds": round(seconds, 3),
                        "timecode": format_seconds(seconds),
                        "image_path": str(image_path),
                        "score": score,
                        "reason": reason,
                        "visual_features": visual_features,
                    }
                )
        prev_frame = frame

    capture.release()

    # --- Assign best keyframe to scenedetect candidates ---
    if detection_method == "scenedetect":
        for candidate in scene_candidates:
            cs = candidate["seconds"]
            best_frame = None
            best_sharpness = -1.0
            # Primary window: within 3x sample_interval after the cut point
            for sf in sampled_frames:
                if sf["seconds"] >= cs and sf["seconds"] < cs + sample_interval * 3:
                    if sf["sharpness"] > best_sharpness:
                        best_sharpness = sf["sharpness"]
                        best_frame = sf
            # Fallback: if no frame found in primary window, find the closest
            # sampled frame by absolute time distance (fixes null keyframe bug)
            if best_frame is None and sampled_frames:
                closest = min(sampled_frames, key=lambda sf: abs(sf["seconds"] - cs))
                best_frame = closest
                notes.append(
                    f"Keyframe fallback for candidate at {format_seconds(cs)}: "
                    f"no frame in primary window, using closest frame at "
                    f"{format_seconds(closest['seconds'])} "
                    f"(distance={abs(closest['seconds'] - cs):.3f}s)"
                )
            if best_frame:
                candidate["image_path"] = best_frame["image_path"]
                candidate["sharpness"] = best_frame.get("sharpness", 0.0)
                candidate["visual_features"] = best_frame.get("visual_features")

    if len(scene_candidates) <= 1:
        notes.append("Very few scene candidates detected; consider manual merging or denser sampling in draft phase")

    draft_blocks = build_draft_blocks(scene_candidates, float(effective_end))

    if not draft_blocks:
        notes.append("ERROR: No draft blocks generated. Scene detection may have failed or video is too short.")
        print("ERROR: No draft blocks could be generated from this video. "
              "Check scene detection settings or try a different --sample-interval.", file=sys.stderr)

    # --- Phase 2.2: ASR ---
    asr_result: dict[str, Any] = {"status": "not-run", "segments": []}
    if args.asr:
        notes.append("Attempting ASR speech recognition...")
        audio_path, audio_err = extract_audio_track(video_path, out_dir, start=effective_start, end=effective_end)
        if audio_path:
            asr_model = args.asr_model if hasattr(args, "asr_model") and args.asr_model else "base"
            segments, asr_err = run_asr_faster_whisper(audio_path, model_size=asr_model)
            if segments is not None:
                # Offset timestamps to absolute video time (audio was clipped to effective range)
                if effective_start > 0.001:
                    for seg in segments:
                        seg["start"] = round(seg["start"] + effective_start, 3)
                        seg["end"] = round(seg["end"] + effective_start, 3)
                        seg["start_time"] = format_seconds(seg["start"])
                        seg["end_time"] = format_seconds(seg["end"])
                asr_result = {"status": "ok", "model": asr_model, "segments": segments}
                notes.append(f"ASR complete, recognized {len(segments)} speech segments (model={asr_model})")
            elif asr_err and "not installed" in asr_err:
                asr_result = {"status": "unavailable", "segments": [], "error": asr_err}
                notes.append(f"faster-whisper unavailable, ASR skipped: {asr_err}")
            else:
                asr_result = {"status": "runtime-failed", "segments": [], "error": asr_err or "unknown"}
                notes.append(f"ASR runtime failed (continuing without speech data): {asr_err}")
        else:
            asr_result = {"status": "no-audio", "segments": [], "error": audio_err or "unknown"}
            notes.append(f"Cannot extract audio track, ASR skipped: {audio_err}")

    # --- Phase 2.3: OCR ---
    ocr_result: dict[str, Any] = {"status": "not-run", "detections": []}
    if args.ocr:
        notes.append("Attempting OCR text recognition...")
        # Run OCR on scene candidate keyframes only (not all frames)
        ocr_frame_paths = []
        for candidate in scene_candidates:
            ip = candidate.get("image_path")
            if ip and Path(ip).exists():
                ocr_frame_paths.append(ip)
        if ocr_frame_paths:
            detections, ocr_err = run_ocr_on_frames(ocr_frame_paths)
            if ocr_err is not None:
                # Engine-level failure (no engine available)
                if "No OCR engine" in ocr_err:
                    ocr_result = {"status": "unavailable", "detections": [], "error": ocr_err}
                    notes.append(f"OCR engine unavailable, skipped: {ocr_err}")
                elif detections:
                    # Partial success — some frames OK, some failed
                    ocr_result = {"status": "partial-ok", "detections": detections, "error": ocr_err}
                    notes.append(f"OCR partial: text detected in {len(detections)} frame(s), but {ocr_err}")
                else:
                    ocr_result = {"status": "runtime-failed", "detections": [], "error": ocr_err}
                    notes.append(f"OCR runtime failed (continuing without text data): {ocr_err}")
            elif detections is not None and len(detections) > 0:
                ocr_result = {"status": "ok", "detections": detections}
                notes.append(f"OCR complete, text detected in {len(detections)} frames")
            else:
                ocr_result = {"status": "ok-no-text", "detections": []}
                notes.append("OCR completed successfully but no on-screen text was detected")
        else:
            ocr_result = {"status": "no-frames", "detections": []}
            notes.append("No keyframes available for OCR")

    # --- Build agent_summary: compact overview for LLM consumption ---
    # This avoids the need to read the full sampled_frames array (which can be
    # 50-200KB for a 10-minute video). LLM should read ONLY agent_summary for
    # Step 3b fill-in, not the full analysis.json.
    KEYFRAME_BATCH_SIZE = 6
    block_overview: list[dict[str, Any]] = []
    all_keyframe_paths: list[str] = []
    for block in draft_blocks:
        kf = block.get("keyframe")
        kf_rel = ""
        if kf:
            try:
                kf_rel = os.path.relpath(kf, str(out_dir)).replace("\\", "/")
            except Exception:
                kf_rel = str(Path(kf).name) if kf else ""
            all_keyframe_paths.append(str(kf))
        block_overview.append({
            "id": block["shot_block"],
            "start": block["start_time"],
            "end": block["end_time"],
            "keyframe": kf_rel,
            "cut_reason": block.get("cut_reason", ""),
            "visual_features": block.get("visual_features"),
        })

    # Group keyframes into batches for efficient LLM viewing
    keyframe_batches: list[list[str]] = []
    for i in range(0, len(all_keyframe_paths), KEYFRAME_BATCH_SIZE):
        keyframe_batches.append(all_keyframe_paths[i:i + KEYFRAME_BATCH_SIZE])

    # Compact ASR summary (only first line of each segment, max 20)
    asr_compact: list[dict[str, str]] = []
    for seg in asr_result.get("segments", [])[:20]:
        asr_compact.append({
            "time": f"{seg.get('start_time', '')} - {seg.get('end_time', '')}",
            "text": seg.get("text", "")[:120],
        })

    # Compact OCR summary
    ocr_compact: list[dict[str, Any]] = []
    for det in ocr_result.get("detections", [])[:15]:
        ocr_compact.append({
            "frame": Path(det.get("frame", "")).name,
            "texts": det.get("texts", [])[:5],
        })

    agent_summary = {
        "_purpose": "Compact overview for LLM fill-in. Read THIS instead of the full analysis.json.",
        "video_duration": video_info.get("duration_timecode", ""),
        "video_resolution": f"{video_info.get('resolution', {}).get('width', '')}x{video_info.get('resolution', {}).get('height', '')}",
        "total_blocks": len(draft_blocks),
        "detection_method": detection_method,
        "blocks": block_overview,
        "keyframe_batches": keyframe_batches,
        "asr_status": asr_result["status"],
        "asr_segments": asr_compact,
        "ocr_status": ocr_result["status"],
        "ocr_detections": ocr_compact,
        "degradation_notes": [n for n in notes if "unavailable" in n.lower() or "failed" in n.lower() or "degraded" in n.lower() or "fallback" in n.lower()],
    }

    analysis = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "video": video_info,
        "agent_summary": agent_summary,
        "analysis_config": {
            "sample_interval_sec": sample_interval,
            "scene_threshold": sd_threshold,
            "content_threshold": sd_content_threshold,
            "detection_method": detection_method,
            "max_samples": max_samples,
            "max_width": int(args.max_width),
            "asr_enabled": bool(args.asr),
            "ocr_enabled": bool(args.ocr),
            "effective_range": {
                "start": round(effective_start, 3),
                "start_time": format_seconds(effective_start),
                "end": round(effective_end, 3),
                "end_time": format_seconds(effective_end),
                "is_clip": effective_start > 0.001 or effective_end < duration - 0.001,
            },
        },
        "scene_candidates": scene_candidates,
        "sampled_frames": sampled_frames,
        "draft_blocks": draft_blocks,
        "asr": asr_result,
        "ocr": ocr_result,
        "notes": notes,
        "degradation": {
            "asr": asr_result["status"],
            "ocr": ocr_result["status"],
            "scene_detection": detection_method,
        },
    }

    analysis_path = out_dir / "analysis.json"
    write_json(analysis_path, analysis)

    # Build structured summary for agent consumption
    generated_files = [str(analysis_path), str(keyframe_dir)]
    if asr_result.get("status") == "ok":
        generated_files.append(str(out_dir / "audio.wav"))

    summary: dict[str, Any] = {
        "status": "ok",
        "stage": "scan-video",
        "output_directory": str(out_dir),
        "generated": generated_files,
        "keyframe_count": len(sampled_frames),
        "scene_candidates": len(scene_candidates),
        "draft_blocks": len(draft_blocks),
        "detection_method": detection_method,
        "warnings": [n for n in notes if "unavailable" in n.lower() or "failed" in n.lower() or "degraded" in n.lower() or "skipped" in n.lower()],
        "notes": notes,
        "degradation": analysis["degradation"],
        "next_recommended_step": f"draft-from-analysis --analysis-json {analysis_path} --output {out_dir / 'cue_sheet.md'} --template production",
    }

    if args.output_format == "json":
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        # Human-readable output
        print(f"Output directory: {out_dir}")
        print(f"  analysis.json : {analysis_path}")
        print(f"  keyframes/    : {keyframe_dir} ({len(sampled_frames)} frames)")
        if asr_result.get("status") == "ok":
            print(f"  audio.wav     : {out_dir / 'audio.wav'}")
        if summary["warnings"]:
            for w in summary["warnings"]:
                print(f"  ⚠ {w}")
        print(f"Next step: draft-from-analysis")

    return 0


def relpath_for_markdown(target: str | None, md_path: Path) -> str:
    if not target:
        return ""
    try:
        return os.path.relpath(target, md_path.parent).replace("\\", "/")
    except Exception:
        return target.replace("\\", "/")


def cmd_draft_from_analysis(args: argparse.Namespace) -> int:
    analysis_path = Path(args.analysis_json)
    if not analysis_path.exists():
        raise FileNotFoundError(f"analysis.json not found: {analysis_path}")

    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    output_path = Path(args.output)
    ensure_parent(output_path)

    video = analysis.get("video", {})
    blocks = analysis.get("draft_blocks", [])
    notes = analysis.get("notes", [])
    asr_data = analysis.get("asr", {})
    ocr_data = analysis.get("ocr", {})
    agent_summary = analysis.get("agent_summary", {})
    template = args.template

    if not blocks:
        print("ERROR: No draft blocks found in analysis.json. "
              "Cannot generate draft. Check scan-video output.", file=sys.stderr)
        return 1

    # --- Template-specific column definitions for the draft table ---
    DRAFT_COLUMNS: dict[str, list[tuple[str, str]]] = {
        "production": [
            ("Shot Block", "shot_block"),
            ("Start", "start_time"),
            ("End", "end_time"),
            ("Keyframe", "_keyframe"),
            ("Shot Size", "_placeholder"),
            ("Angle/Lens", "_placeholder"),
            ("Motion", "_placeholder"),
            ("Scene", "_placeholder"),
            ("Mood", "_placeholder"),
            ("Location", "_placeholder"),
            ("Characters", "_placeholder"),
            ("Event", "_placeholder"),
            ("Dialogue", "_placeholder"),
            ("Music Note", "_placeholder"),
            ("Director Note", "_placeholder"),
            ("Cut Reason", "cut_reason"),
            ("Confidence", "_confidence"),
            ("Needs Confirmation", "_placeholder"),
        ],
        "music-director": [
            ("Shot Block", "shot_block"),
            ("Start", "start_time"),
            ("End", "end_time"),
            ("Mood", "_placeholder"),
            ("Event", "_placeholder"),
            ("Dialogue", "_placeholder"),
            ("Music Note", "_placeholder"),
            ("Rhythm Change", "_placeholder"),
            ("Instrumentation", "_placeholder"),
            ("Dynamics", "_placeholder"),
            ("Cut Reason", "cut_reason"),
            ("Confidence", "_confidence"),
            ("Needs Confirmation", "_placeholder"),
        ],
        "script": [
            ("Shot Block", "shot_block"),
            ("Start", "start_time"),
            ("End", "end_time"),
            ("Scene", "_placeholder"),
            ("Location", "_placeholder"),
            ("Characters", "_placeholder"),
            ("Event", "_placeholder"),
            ("Dialogue", "_placeholder"),
            ("Cut Reason", "cut_reason"),
            ("Confidence", "_confidence"),
            ("Needs Confirmation", "_placeholder"),
        ],
    }

    draft_cols = DRAFT_COLUMNS.get(template, DRAFT_COLUMNS["production"])

    lines: list[str] = []
    lines.append(f"# Cue Sheet Draft ({template})")
    lines.append("")

    # --- Video info ---
    lines.append("## Video Info")
    lines.append("")
    lines.append("| Field | Value |")
    lines.append("|---|---|")
    lines.append(f"| Source | `{video.get('source_path', '')}` |")
    lines.append(f"| Duration | {video.get('duration_timecode', '')} |")
    res = video.get("resolution", {})
    lines.append(f"| Resolution | {res.get('width', '')}x{res.get('height', '')} |")
    lines.append(f"| FPS | {video.get('fps', '')} |")
    lines.append(f"| Audio tracks | {video.get('audio_tracks', '')} |")
    eff_range = analysis.get("analysis_config", {}).get("effective_range", {})
    if eff_range.get("is_clip"):
        lines.append(f"| **Analysis range** | {eff_range.get('start_time', '')} — {eff_range.get('end_time', '')} (clip-only) |")
    lines.append("")

    # --- Keyframe batch plan (for efficient LLM viewing) ---
    keyframe_batches = agent_summary.get("keyframe_batches", [])
    if not keyframe_batches:
        # Build from blocks if agent_summary not available (backward compat)
        KEYFRAME_BATCH_SIZE = 6
        all_kf = [b.get("keyframe") for b in blocks if b.get("keyframe")]
        keyframe_batches = [all_kf[i:i + KEYFRAME_BATCH_SIZE] for i in range(0, len(all_kf), KEYFRAME_BATCH_SIZE)]

    if keyframe_batches:
        lines.append("## Keyframe Batches for Fill-in")
        lines.append("")
        lines.append("> **Agent instruction**: View keyframes in these batches (not one-by-one).")
        lines.append("> After viewing each batch, fill in ALL corresponding blocks in the JSON fill-in file.")
        lines.append("")
        for batch_idx, batch in enumerate(keyframe_batches, start=1):
            paths_display = ", ".join(Path(p).name for p in batch)
            block_ids = []
            for b in blocks:
                kf = b.get("keyframe")
                if kf and str(kf) in batch:
                    block_ids.append(b["shot_block"])
            lines.append(f"- **Batch {batch_idx}** (blocks {', '.join(block_ids)}): `{paths_display}`")
        lines.append("")

    # --- Template-differentiated candidate segments ---
    lines.append(f"## Candidate Segments ({template})")
    lines.append("")
    header = "| " + " | ".join(col[0] for col in draft_cols) + " |"
    separator = "|" + "|".join(["---"] * len(draft_cols)) + "|"
    lines.append(header)
    lines.append(separator)

    for block in blocks:
        cells: list[str] = []
        for col_label, col_key in draft_cols:
            if col_key == "shot_block":
                cells.append(block.get("shot_block", ""))
            elif col_key == "start_time":
                cells.append(block.get("start_time", ""))
            elif col_key == "end_time":
                cells.append(block.get("end_time", ""))
            elif col_key == "_keyframe":
                img_rel = relpath_for_markdown(block.get("keyframe"), output_path)
                cells.append(f"![kf]({img_rel})" if img_rel else "*(no keyframe)*")
            elif col_key == "cut_reason":
                cells.append(block.get("cut_reason", ""))
            elif col_key == "_confidence":
                score = block.get("candidate_score")
                cells.append("high" if isinstance(score, (int, float)) and score >= 0.8 else "medium")
            elif col_key == "_placeholder":
                cells.append("*(pending)*")
        lines.append("| " + " | ".join(cells) + " |")
    lines.append("")

    # --- Template-specific fill-in guidance ---
    lines.append("## Fill-in Guidance")
    lines.append("")
    lines.append("> **IMPORTANT**: Use the JSON fill-in file (`draft_fill.json`) instead of editing this Markdown table.")
    lines.append("> The JSON file was generated alongside this draft. Fill in the empty fields there,")
    lines.append("> then the final export steps will consume the JSON directly.")
    lines.append("")
    if template == "production":
        lines.append("For each block, fill in these fields in `draft_fill.json`:")
        lines.append("- **shot_size**: WS / MS / CU / EWS / ECU")
        lines.append("- **angle_or_lens**: front / OTS / low angle / high angle / eye-level")
        lines.append("- **motion**: static / push-in / pull-out / pan / tracking / handheld")
        lines.append("- **scene**: location/setup name (use `temp: xxx` if unconfirmed)")
        lines.append("- **mood**: emotional tone of the block")
        lines.append("- **location**: where this takes place")
        lines.append("- **characters**: who is present (use `temp: xxx` if unconfirmed)")
        lines.append("- **event**: what happens in this block")
        lines.append("- **important_dialogue**: key dialogue lines")
        lines.append("- **music_note**: music entry/exit/change suggestions")
        lines.append("- **director_note**: camera language, cross-dept coordination points")
        lines.append("- **confidence**: e.g. `segment=high; names=low`")
        lines.append("- **needs_confirmation**: items that need user confirmation")
    elif template == "music-director":
        lines.append("For each block, fill in these fields in `draft_fill.json`:")
        lines.append("- **mood**: emotional tone and progression within the block")
        lines.append("- **event**: what happens (context for music decisions)")
        lines.append("- **important_dialogue**: key lines that music must accommodate")
        lines.append("- **music_note**: entry/exit/change points, overall direction")
        lines.append("- **rhythm_change**: BPM shifts, pulse changes, rhythmic events")
        lines.append("- **instrumentation**: specific instrument/texture suggestions")
        lines.append("- **dynamics**: pp/mp/mf/f/ff progression")
        lines.append("- **confidence**: e.g. `mood=medium; dialogue=low`")
        lines.append("- **needs_confirmation**: items that need user confirmation")
    elif template == "script":
        lines.append("For each block, fill in these fields in `draft_fill.json`:")
        lines.append("- **scene**: scene name or setup (use `temp: xxx` if unconfirmed)")
        lines.append("- **location**: where this takes place")
        lines.append("- **characters**: who is present")
        lines.append("- **event**: what happens — focus on story, not camera")
        lines.append("- **important_dialogue**: key lines / narration / inner monologue")
        lines.append("- **confidence**: e.g. `segment=high; dialogue=low`")
        lines.append("- **needs_confirmation**: items that need user confirmation")
    lines.append("")

    # --- ASR Highlights ---
    asr_segments = asr_data.get("segments", [])
    if asr_segments:
        lines.append("## ASR Highlights")
        lines.append("")
        lines.append("| Start | End | Text |")
        lines.append("|---|---|---|")
        for seg in asr_segments[:30]:
            text = seg.get("text", "").replace("|", "\\|").replace("\n", " ")
            lines.append(f"| {seg.get('start_time', '')} | {seg.get('end_time', '')} | {text} |")
        if len(asr_segments) > 30:
            lines.append(f"| ... | ... | ({len(asr_segments) - 30} more segments in analysis.json) |")
        lines.append("")
    elif asr_data.get("status") == "not-run":
        lines.append("## ASR Highlights")
        lines.append("")
        lines.append("*ASR was not enabled for this scan. Re-run with `--asr` to include speech recognition.*")
        lines.append("")
    else:
        asr_status = asr_data.get("status", "unknown")
        asr_error = asr_data.get("error", "")
        lines.append("## ASR Highlights")
        lines.append("")
        lines.append(f"*ASR status: {asr_status}.* {asr_error}")
        lines.append("")

    # --- OCR Highlights ---
    ocr_detections = ocr_data.get("detections", [])
    if ocr_detections:
        lines.append("## OCR Highlights")
        lines.append("")
        lines.append("| Frame | Detected Text |")
        lines.append("|---|---|")
        for det in ocr_detections:
            frame_name = Path(det.get("frame", "")).name
            texts = "; ".join(det.get("texts", []))
            lines.append(f"| {frame_name} | {texts} |")
        lines.append("")
    elif ocr_data.get("status") == "not-run":
        lines.append("## OCR Highlights")
        lines.append("")
        lines.append("*OCR was not enabled for this scan. Re-run with `--ocr` to include text detection.*")
        lines.append("")
    elif ocr_data.get("status") == "ok-no-text":
        lines.append("## OCR Highlights")
        lines.append("")
        lines.append("*OCR completed successfully but no on-screen text was detected in the analyzed frames.*")
        lines.append("")
    else:
        ocr_status = ocr_data.get("status", "unknown")
        ocr_error = ocr_data.get("error", "")
        lines.append("## OCR Highlights")
        lines.append("")
        lines.append(f"*OCR status: {ocr_status}.* {ocr_error}")
        lines.append("")

    # --- Character / Scene / Prop Summary (LLM fills based on keyframes) ---
    lines.append("## Character Summary")
    lines.append("")
    lines.append("*(Fill based on keyframe analysis — list all identified characters with brief visual description, "
                 "role in the analyzed segment, and evidence from specific blocks.)*")
    lines.append("")
    lines.append("| Temp Name | Visual Description | Appears in Blocks | Role / Notes |")
    lines.append("|---|---|---|---|")
    lines.append("| temp: Character-A | *(describe appearance)* | *(list blocks)* | *(role or behavior)* |")
    lines.append("")

    lines.append("## Scene / Setup Summary")
    lines.append("")
    lines.append("*(Fill based on keyframe analysis — list all identified locations or setups.)*")
    lines.append("")
    lines.append("| Temp Name | Space Description | Appears in Blocks | Notes |")
    lines.append("|---|---|---|---|")
    lines.append("| temp: Scene-A | *(describe space)* | *(list blocks)* | *(lighting, mood, notable features)* |")
    lines.append("")

    lines.append("## Prop Summary")
    lines.append("")
    lines.append("*(Fill if any key props are identified in close-ups, handoffs, or repeated appearances.)*")
    lines.append("")
    lines.append("| Temp Name | Description | Appears in Blocks | Importance |")
    lines.append("|---|---|---|---|")
    lines.append("| temp: Prop-A | *(describe prop)* | *(list blocks)* | *(key-prop / background)* |")
    lines.append("")

    # --- Naming confirmation tables ---
    lines.append("## Naming Confirmation Tables")
    lines.append("")
    lines.append("### Characters")
    lines.append("")
    lines.append("| temporary_name | evidence | confidence | confirmed_name | status |")
    lines.append("|---|---|---|---|---|")
    lines.append("| temp: Character-A | Fill based on keyframe & behavior | low |  | pending |")
    lines.append("")
    lines.append("### Scenes / Setups")
    lines.append("")
    lines.append("| temporary_setup | space_note | confidence | confirmed_setup | status |")
    lines.append("|---|---|---|---|---|")
    lines.append("| temp: Scene-A | Fill based on space & establishing shots | low |  | pending |")
    lines.append("")
    lines.append("### Props")
    lines.append("")
    lines.append("| temporary_prop | importance | evidence | confirmed_prop | status |")
    lines.append("|---|---|---|---|---|")
    lines.append("| temp: Prop-A | key-prop? | Fill based on close-ups & repeated appearances |  | pending |")
    lines.append("")

    # --- Pending questions ---
    lines.append("## Pending Questions")
    lines.append("")
    lines.append("- Do characters have official names?")
    lines.append("- Do scenes have internal project setup names?")
    lines.append("- Do key props have standardized names?")
    lines.append(f"- Proceed with `{template}` template for final?")
    lines.append("- Export final Excel?")
    lines.append("")

    # --- Notes ---
    lines.append("## Notes & Fallback Info")
    lines.append("")
    if not notes:
        lines.append("- No additional notes.")
    else:
        for note in notes:
            lines.append(f"- {note}")
    lines.append("")

    # --- Next steps ---
    lines.append("## Next Steps")
    lines.append("")
    lines.append("1. View keyframes in batches listed above.")
    lines.append("2. Fill in empty fields in `draft_fill.json` (the JSON fill-in file generated alongside this draft).")
    lines.append("3. Confirm character / scene / prop names (or keep temp markers).")
    lines.append("4. Merge blocks using `merge-blocks` if needed.")
    lines.append("5. Generate `final_cues.json` (or use `build-final-skeleton`).")
    lines.append("6. Validate with `validate-cue-json`.")
    lines.append("7. Export with `build-xlsx` and `export-md`.")
    lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")

    # --- Generate JSON fill-in file (draft_fill.json) ---
    # This is the primary file the LLM should edit instead of the Markdown table.
    # Pre-populated fields (shot_block, start_time, end_time, keyframe) are filled in.
    # Data-derived fields (important_dialogue, confidence, needs_confirmation, etc.)
    # are pre-filled from ASR/OCR/analysis data. Remaining fields are empty for LLM.
    fill_in_path = output_path.parent / "draft_fill.json"
    columns = TEMPLATE_COLUMNS[template]

    # Build lookup structures for pre-fill
    asr_segments = asr_data.get("segments", [])
    ocr_detections = ocr_data.get("detections", [])
    ocr_by_frame: dict[str, list[str]] = {}
    for det in ocr_detections:
        frame_path = det.get("frame", "")
        ocr_by_frame[frame_path] = det.get("texts", [])
        # Also index by filename for cross-platform path matching
        if frame_path:
            ocr_by_frame[Path(frame_path).name] = det.get("texts", [])

    detection_method = analysis.get("analysis_config", {}).get("detection_method", "histogram")
    asr_status = asr_data.get("status", "not-run")

    has_prefill = False
    fill_rows: list[dict[str, Any]] = []
    for block in blocks:
        row: dict[str, Any] = {}
        kf_raw = block.get("keyframe")
        kf_rel = ""
        if kf_raw:
            try:
                kf_rel = os.path.relpath(kf_raw, str(output_path.parent)).replace("\\", "/")
            except Exception:
                kf_rel = str(Path(kf_raw).name) if kf_raw else ""

        block_start = block.get("start_seconds", 0.0)
        block_end = block.get("end_seconds", 0.0)
        vf = block.get("visual_features") or {}

        for col in columns:
            if col == "shot_block":
                row[col] = block.get("shot_block", "")
            elif col == "start_time":
                row[col] = block.get("start_time", "")
            elif col == "end_time":
                row[col] = block.get("end_time", "")
            elif col == "keyframe":
                row[col] = kf_rel

            # --- Pre-fill: important_dialogue from ASR ---
            elif col == "important_dialogue" and asr_segments:
                overlapping = [
                    s for s in asr_segments
                    if s.get("start", 0) < block_end and s.get("end", 0) > block_start
                ]
                if overlapping:
                    row[col] = "; ".join(
                        f"[{s.get('start_time', '')}] {s.get('text', '')}" for s in overlapping
                    )
                    has_prefill = True
                else:
                    row[col] = ""

            # --- Pre-fill: confidence from structural data ---
            elif col == "confidence":
                parts = []
                score = block.get("candidate_score")
                if detection_method == "scenedetect":
                    parts.append("segment=high")
                elif isinstance(score, (int, float)) and score >= 0.8:
                    parts.append("segment=high")
                else:
                    parts.append("segment=medium")
                # ASR coverage
                if asr_status == "ok":
                    asr_overlap = any(
                        s.get("start", 0) < block_end and s.get("end", 0) > block_start
                        for s in asr_segments
                    )
                    parts.append("dialogue=high" if asr_overlap else "dialogue=low")
                elif asr_status != "not-run":
                    parts.append("dialogue=degraded")
                parts.append("names=low")
                row[col] = "; ".join(parts)
                has_prefill = True

            # --- Pre-fill: needs_confirmation (always needed) ---
            elif col == "needs_confirmation":
                row[col] = "character names; scene names"
                has_prefill = True

            # --- Pre-fill: mood with visual feature hints ---
            elif col == "mood" and vf:
                # Provide objective hints, not a final mood label
                hints = []
                if vf.get("tone"):
                    hints.append(f"{vf['tone']} tones")
                if vf.get("color_temp") and vf["color_temp"] != "neutral":
                    hints.append(f"{vf['color_temp']} color")
                if vf.get("contrast", 0) > 70:
                    hints.append("high contrast")
                elif vf.get("contrast", 0) < 30:
                    hints.append("low contrast")
                if vf.get("saturation", 0) < 40:
                    hints.append("desaturated")
                elif vf.get("saturation", 0) > 150:
                    hints.append("vivid")
                if hints:
                    row[col] = f"[visual: {', '.join(hints)}] "
                    has_prefill = True
                else:
                    row[col] = ""

            # --- Default: empty for LLM to fill ---
            else:
                # Check OCR for director_note / event enrichment
                if col in ("director_note", "event") and kf_raw:
                    kf_name = Path(kf_raw).name if kf_raw else ""
                    ocr_texts = ocr_by_frame.get(str(kf_raw), []) or ocr_by_frame.get(kf_name, [])
                    if ocr_texts:
                        row[col] = f"[OCR detected: {'; '.join(ocr_texts[:3])}] "
                        has_prefill = True
                    else:
                        row[col] = ""
                else:
                    row[col] = ""
        fill_rows.append(row)

    fill_status = "partial" if has_prefill else "pending"
    fill_data = {
        "_instructions": (
            "Fill in the empty string fields for each block based on keyframe analysis. "
            "Fields starting with '[visual: ...]' or '[OCR detected: ...]' contain auto-generated hints — "
            "incorporate or replace them with your analysis. "
            "Do NOT modify shot_block, start_time, end_time, or keyframe. "
            "Use 'temp: xxx' for unconfirmed names. "
            "After filling, this file can be used directly as input to build-final-skeleton or validate-cue-json."
        ),
        "template": template,
        "video_title": Path(video.get("source_path", "untitled")).stem if video.get("source_path") else "untitled",
        "source_path": video.get("source_path", ""),
        "fill_status": fill_status,
        "rows": fill_rows,
    }
    write_json(fill_in_path, fill_data)

    print(str(output_path))
    print(f"JSON fill-in file: {fill_in_path}")
    return 0


def resolve_keyframe_path(base_dir: Path | None, value: str | None) -> Path | None:
    if not value:
        return None
    candidate = Path(value)
    if candidate.is_absolute():
        return candidate
    if base_dir:
        return (base_dir / candidate).resolve()
    return candidate.resolve()


def scale_dimensions(width: int, height: int, max_width: int, max_height: int) -> tuple[int, int]:
    if width <= 0 or height <= 0:
        return max_width, max_height
    ratio = min(max_width / width, max_height / height, 1.0)
    return max(int(width * ratio), 1), max(int(height * ratio), 1)


def cmd_build_xlsx(args: argparse.Namespace) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from PIL import Image as PILImage
    except Exception as exc:
        raise RuntimeError(f"Modules required for Excel export unavailable: {exc}") from exc

    cue_json = Path(args.cue_json)
    if not cue_json.exists():
        raise FileNotFoundError(f"Cue JSON not found: {cue_json}")

    payload = json.loads(cue_json.read_text(encoding="utf-8"))
    template = args.template or payload.get("template") or "production"
    if template not in TEMPLATE_COLUMNS:
        raise ValueError(f"Unknown template: {template}")

    base_dir = Path(args.base_dir).resolve() if args.base_dir else cue_json.parent.resolve()
    rows = payload.get("rows", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Cue Sheet"
    meta = wb.create_sheet("Meta")

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    columns = TEMPLATE_COLUMNS[template]
    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment
        ws.column_dimensions[cell.column_letter].width = DEFAULT_COLUMN_WIDTHS.get(column_name, 18)

    keyframe_col_index = columns.index("keyframe") + 1 if "keyframe" in columns else None

    # --- Delivery completeness tracking ---
    embedded_keyframes = 0
    missing_keyframes: list[str] = []
    empty_recommended_fields = 0
    RECOMMENDED_FIELDS: dict[str, list[str]] = {
        "production": ["scene", "event", "shot_size", "mood", "characters"],
        "music-director": ["mood", "event", "music_note", "rhythm_change", "dynamics"],
        "script": ["scene", "event", "characters", "location"],
    }
    rec_fields = RECOMMENDED_FIELDS.get(template, [])

    for row_idx, item in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(columns, start=1):
            value = item.get(column_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value if column_name != "keyframe" else "")
            cell.alignment = wrap_alignment

        # Track empty recommended fields
        for rf in rec_fields:
            if rf in set(columns) and not str(item.get(rf, "")).strip():
                empty_recommended_fields += 1

        if keyframe_col_index is not None:
            keyframe_value = item.get("keyframe")
            keyframe_path = resolve_keyframe_path(base_dir, keyframe_value)
            if keyframe_path and keyframe_path.exists():
                with PILImage.open(keyframe_path) as image:
                    width, height = image.size
                scaled_w, scaled_h = scale_dimensions(width, height, args.image_max_width, args.image_max_height)
                # Resize image before embedding to reduce xlsx size and speed up generation
                with PILImage.open(keyframe_path) as img:
                    thumb = img.resize((scaled_w, scaled_h), PILImage.LANCZOS)
                    thumb_path = keyframe_path.parent / f".thumb_{keyframe_path.name}"
                    thumb.save(str(thumb_path), "JPEG", quality=85)
                xl_image = XLImage(str(thumb_path))
                xl_image.width = scaled_w
                xl_image.height = scaled_h
                anchor_cell = ws.cell(row=row_idx, column=keyframe_col_index)
                ws.add_image(xl_image, anchor_cell.coordinate)
                anchor_cell.value = keyframe_path.name
                # ~0.78 converts pixels to approximate Excel row-height points
                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, scaled_h * 0.78)
                embedded_keyframes += 1
            else:
                block_label = item.get("shot_block", f"row{row_idx - 1}")
                missing_keyframes.append(block_label)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta_rows = [
        ("template", template),
        ("video_title", payload.get("video_title", "")),
        ("source_path", payload.get("source_path", "")),
        ("generated_at", payload.get("generated_at", dt.datetime.now().isoformat(timespec="seconds"))),
        ("row_count", len(rows)),
    ]
    for index, (key, value) in enumerate(meta_rows, start=1):
        meta.cell(row=index, column=1, value=key)
        meta.cell(row=index, column=2, value=value)

    output_path = Path(args.output)
    ensure_parent(output_path)
    wb.save(output_path)

    # Clean up temporary thumbnail files
    if keyframe_col_index is not None:
        for item in rows:
            kf_value = item.get("keyframe")
            kf_path = resolve_keyframe_path(base_dir, kf_value)
            if kf_path:
                thumb = kf_path.parent / f".thumb_{kf_path.name}"
                if thumb.exists():
                    try:
                        thumb.unlink()
                    except OSError:
                        pass

    print(str(output_path))

    # --- Delivery completeness summary ---
    print("--- delivery summary ---")
    print(f"  rows exported: {len(rows)}")
    if not rows:
        print("  WARNING: no rows exported — cue sheet is empty")
    if keyframe_col_index is not None:
        print(f"  embedded keyframes: {embedded_keyframes}/{len(rows)}")
        if missing_keyframes:
            print(f"  WARNING: missing keyframes for blocks: {', '.join(missing_keyframes)}")
    if empty_recommended_fields > 0:
        print(f"  WARNING: {empty_recommended_fields} empty recommended field(s) across all rows")
    delivery_ready = len(rows) > 0 and (not missing_keyframes or keyframe_col_index is None) and empty_recommended_fields == 0
    print(f"  delivery_ready: {'YES' if delivery_ready else 'NO — review warnings above'}")
    return 0


def cmd_validate_cue_json(args: argparse.Namespace) -> int:
    cue_json = Path(args.cue_json)
    if not cue_json.exists():
        raise FileNotFoundError(f"Cue JSON not found: {cue_json}")

    payload = json.loads(cue_json.read_text(encoding="utf-8"))
    template = args.template or payload.get("template") or "production"
    if template not in TEMPLATE_COLUMNS:
        raise ValueError(f"Unknown template: {template}")

    rows = payload.get("rows", [])
    expected_columns = set(TEMPLATE_COLUMNS[template])
    errors: list[str] = []
    warnings: list[str] = []

    if not rows:
        errors.append("rows is empty; at least one shot block is required.")

    prev_end: float | None = None
    seen_blocks: set[str] = set()

    for idx, row in enumerate(rows):
        label = row.get("shot_block", f"row[{idx}]")

        if label in seen_blocks:
            errors.append(f"{label}: duplicate shot_block ID.")
        seen_blocks.add(label)

        for required_field in ("shot_block", "start_time", "end_time"):
            if not row.get(required_field):
                errors.append(f"{label}: missing required field '{required_field}'.")

        start_text = row.get("start_time", "")
        end_text = row.get("end_time", "")
        start_sec: float | None = None
        end_sec: float | None = None
        try:
            if start_text:
                start_sec = seconds_from_timecode(start_text)
        except Exception:
            errors.append(f"{label}: invalid start_time format '{start_text}', expected HH:MM:SS.mmm.")
        try:
            if end_text:
                end_sec = seconds_from_timecode(end_text)
        except Exception:
            errors.append(f"{label}: invalid end_time format '{end_text}', expected HH:MM:SS.mmm.")

        if start_sec is not None and end_sec is not None:
            if end_sec < start_sec:
                errors.append(f"{label}: end_time ({end_text}) is before start_time ({start_text}).")

        if prev_end is not None and start_sec is not None:
            gap = abs(start_sec - prev_end)
            if gap > 0.1:
                warnings.append(f"{label}: {gap:.3f}s gap from previous block end.")
            if start_sec < prev_end - 0.05:
                warnings.append(f"{label}: {prev_end - start_sec:.3f}s overlap with previous block.")

        if end_sec is not None:
            prev_end = end_sec

        row_keys = set(row.keys())
        extra_keys = row_keys - expected_columns
        if extra_keys:
            warnings.append(f"{label}: fields not defined in template '{template}': {', '.join(sorted(extra_keys))}.")

        naming_fields = ("scene", "characters", "location")
        needs_conf = str(row.get("needs_confirmation", "")).strip()
        for nf in naming_fields:
            value = str(row.get(nf, ""))
            if "temp:" in value.lower() and not needs_conf:
                warnings.append(f"{label}: '{nf}' contains temp name but needs_confirmation is empty.")

        # Per-template recommended field completeness
        RECOMMENDED_FIELDS: dict[str, list[str]] = {
            "production": ["scene", "event", "shot_size", "mood", "characters"],
            "music-director": ["mood", "event", "music_note", "rhythm_change", "dynamics"],
            "script": ["scene", "event", "characters", "location"],
        }
        for rec_field in RECOMMENDED_FIELDS.get(template, []):
            if rec_field in expected_columns and not str(row.get(rec_field, "")).strip():
                warnings.append(f"{label}: recommended field '{rec_field}' is empty for template '{template}'.")

    # --- Optional: keyframe file existence check ---
    if hasattr(args, "check_files") and args.check_files and "keyframe" in expected_columns:
        base_dir = Path(args.base_dir).resolve() if hasattr(args, "base_dir") and args.base_dir else cue_json.parent.resolve()
        for row in rows:
            label = row.get("shot_block", "?")
            kf_value = row.get("keyframe", "")
            if kf_value:
                kf_path = resolve_keyframe_path(base_dir, kf_value)
                if kf_path and not kf_path.exists():
                    warnings.append(f"{label}: keyframe file not found: {kf_path}")

    # Delivery readiness: no warnings about empty recommended fields, temp-name inconsistencies, or missing keyframes
    delivery_gaps = [w for w in warnings if "recommended field" in w or "temp name" in w.lower() or "needs_confirmation is empty" in w or "keyframe file not found" in w]
    delivery_ready = len(errors) == 0 and len(delivery_gaps) == 0

    report = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "template": template,
        "total_rows": len(rows),
        "errors": errors,
        "warnings": warnings,
        "valid": len(errors) == 0,
        "delivery_ready": delivery_ready,
        "delivery_gaps": delivery_gaps,
    }

    if args.report_out:
        write_json(Path(args.report_out), report)

    if args.output_format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(f"=== validate-cue-json ({template}) ===")
        print(f"Rows: {len(rows)}")
        if errors:
            print(f"Errors ({len(errors)}):")
            for e in errors:
                print(f"  ✗ {e}")
        if warnings:
            print(f"Warnings ({len(warnings)}):")
            for w in warnings:
                print(f"  ⚠ {w}")
        if not errors and not warnings:
            print("  ✓ No errors, no warnings.")
        print(f"Valid: {'YES' if report['valid'] else 'NO'}")
        print(f"Delivery ready: {'YES' if delivery_ready else 'NO'}")
        if delivery_gaps:
            print(f"Delivery gaps ({len(delivery_gaps)}):")
            for g in delivery_gaps:
                print(f"  → {g}")

    return 0 if report["valid"] else 1


def apply_naming_to_text(text: str, mapping: dict[str, str]) -> str:
    """Apply naming replacements, longest keys first to avoid substring collisions."""
    result = text
    for old_name, new_name in sorted(mapping.items(), key=lambda kv: len(kv[0]), reverse=True):
        result = result.replace(old_name, new_name)
    return result


NAMING_REPLACE_FIELDS = {"scene", "characters", "location", "event", "important_dialogue", "needs_confirmation", "director_note", "music_note"}


def apply_naming_to_json_structured(payload: dict[str, Any], mappings: dict[str, str]) -> tuple[dict[str, Any], int]:
    """Apply naming overrides only to whitelisted fields in rows. Returns (new_payload, change_count).
    Replacements are applied longest-key-first to avoid substring collisions."""
    import copy
    result = copy.deepcopy(payload)
    changes = 0
    sorted_mappings = sorted(mappings.items(), key=lambda kv: len(kv[0]), reverse=True)
    for row in result.get("rows", []):
        for field in NAMING_REPLACE_FIELDS:
            value = row.get(field)
            if not isinstance(value, str):
                continue
            new_value = value
            for old_name, new_name in sorted_mappings:
                new_value = new_value.replace(old_name, new_name)
            if new_value != value:
                row[field] = new_value
                changes += 1
    return result, changes


def cmd_apply_naming(args: argparse.Namespace) -> int:
    overrides_path = Path(args.overrides)
    if not overrides_path.exists():
        raise FileNotFoundError(f"Naming overrides file not found: {overrides_path}")

    overrides = json.loads(overrides_path.read_text(encoding="utf-8"))
    all_mappings: dict[str, str] = {}
    for category in ("characters", "scenes", "props"):
        mappings = overrides.get(category, {})
        all_mappings.update(mappings)

    if not all_mappings:
        print("Naming overrides empty, no replacements needed.")
        return 0

    if not args.cue_json and not args.md:
        print("ERROR: At least one of --cue-json or --md must be specified.", file=sys.stderr)
        return 1

    dry_run = bool(args.dry_run)
    replaced_count = 0
    changes_detail: list[dict[str, Any]] = []

    if args.cue_json:
        cue_path = Path(args.cue_json)
        if not cue_path.exists():
            raise FileNotFoundError(f"Cue JSON not found: {cue_path}")
        payload = json.loads(cue_path.read_text(encoding="utf-8"))
        new_payload, change_count = apply_naming_to_json_structured(payload, all_mappings)
        changes_detail.append({"file": str(cue_path), "type": "json", "field_changes": change_count})
        if change_count > 0:
            if not dry_run:
                out_path = Path(args.output) if args.output else cue_path
                write_json(out_path, new_payload)
                print(f"Updated ({change_count} field changes): {out_path}")
            else:
                print(f"[dry-run] Would update {change_count} fields in: {cue_path}")
            replaced_count += 1
        else:
            print(f"No changes: {cue_path}")

    if args.md:
        md_path = Path(args.md)
        if not md_path.exists():
            raise FileNotFoundError(f"Markdown not found: {md_path}")
        raw = md_path.read_text(encoding="utf-8")
        new_raw = apply_naming_to_text(raw, all_mappings)
        md_changes = 1 if raw != new_raw else 0
        changes_detail.append({"file": str(md_path), "type": "md", "changed": bool(md_changes)})
        if md_changes:
            if not dry_run:
                md_path.write_text(new_raw, encoding="utf-8")
                print(f"Updated: {md_path}")
            else:
                print(f"[dry-run] Would update: {md_path}")
            replaced_count += 1
        else:
            print(f"No changes: {md_path}")

    report = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "dry_run": dry_run,
        "mappings_count": len(all_mappings),
        "mappings": all_mappings,
        "files_processed": changes_detail,
        "files_updated": replaced_count,
    }

    if args.report_out:
        write_json(Path(args.report_out), report)

    mode = "[dry-run] " if dry_run else ""
    print(f"{mode}Total {len(all_mappings)} mappings, {replaced_count} file(s) {'would be ' if dry_run else ''}updated.")
    return 0


# ---------------------------------------------------------------------------
# Naming table derivation
# ---------------------------------------------------------------------------

# Fields that may contain temp: markers, grouped by naming category
NAMING_CATEGORY_FIELDS: dict[str, list[str]] = {
    "characters": ["characters"],
    "scenes": ["scene", "location"],
    "props": ["event", "important_dialogue", "director_note", "music_note"],
}

import re as _re
_TEMP_MARKER_RE = _re.compile(r"temp:\s*[A-Za-z0-9\u4e00-\u9fff][\w\u4e00-\u9fff\-]*(?:\s+[\w\u4e00-\u9fff\-]+)*", _re.UNICODE)


def extract_temp_markers(text: str) -> list[str]:
    """Extract all 'temp: XYZ' markers from a text string."""
    if not text or "temp:" not in text.lower():
        return []
    return _TEMP_MARKER_RE.findall(text)


def derive_naming_tables_from_rows(
    rows: list[dict[str, Any]],
    template: str = "production",
) -> dict[str, list[dict[str, Any]]]:
    """Scan rows for temp: markers, deduplicate, and aggregate block references.

    Returns: {"characters": [...], "scenes": [...], "props": [...]}
    Each entry: {"temporary_name": ..., "appears_in_blocks": [...], "evidence": ..., "confidence": "low"}
    """
    # category -> { marker_text -> set of block IDs }
    category_map: dict[str, dict[str, set[str]]] = {
        "characters": {},
        "scenes": {},
        "props": {},
    }

    for row in rows:
        block_id = row.get("shot_block", "?")
        for category, fields in NAMING_CATEGORY_FIELDS.items():
            for field in fields:
                value = row.get(field, "")
                if not value:
                    continue
                markers = extract_temp_markers(value)
                for marker in markers:
                    normalized = marker.strip()
                    if normalized not in category_map[category]:
                        category_map[category][normalized] = set()
                    category_map[category][normalized].add(block_id)

    # Build structured output
    result: dict[str, list[dict[str, Any]]] = {}
    for category in ("characters", "scenes", "props"):
        entries: list[dict[str, Any]] = []
        for marker, block_ids in sorted(category_map[category].items()):
            sorted_blocks = sorted(block_ids, key=lambda b: (len(b), b))
            entries.append({
                "temporary_name": marker,
                "appears_in_blocks": sorted_blocks,
                "evidence": f"Found in {len(sorted_blocks)} block(s): {', '.join(sorted_blocks)}",
                "confidence": "low",
                "confirmed_name": "",
                "status": "pending",
            })
        result[category] = entries

    return result


def format_naming_tables_md(tables: dict[str, list[dict[str, Any]]]) -> str:
    """Format naming tables as Markdown sections."""
    lines: list[str] = []

    lines.append("## Naming Confirmation Tables")
    lines.append("")

    lines.append("### Characters")
    lines.append("")
    lines.append("| temporary_name | appears_in_blocks | evidence | confidence | confirmed_name | status |")
    lines.append("|---|---|---|---|---|---|")
    for entry in tables.get("characters", []):
        lines.append(
            f"| {entry['temporary_name']} "
            f"| {', '.join(entry['appears_in_blocks'])} "
            f"| {entry['evidence']} "
            f"| {entry['confidence']} "
            f"| {entry['confirmed_name']} "
            f"| {entry['status']} |"
        )
    if not tables.get("characters"):
        lines.append("| *(none detected)* | | | | | |")
    lines.append("")

    lines.append("### Scenes / Setups")
    lines.append("")
    lines.append("| temporary_name | appears_in_blocks | evidence | confidence | confirmed_name | status |")
    lines.append("|---|---|---|---|---|---|")
    for entry in tables.get("scenes", []):
        lines.append(
            f"| {entry['temporary_name']} "
            f"| {', '.join(entry['appears_in_blocks'])} "
            f"| {entry['evidence']} "
            f"| {entry['confidence']} "
            f"| {entry['confirmed_name']} "
            f"| {entry['status']} |"
        )
    if not tables.get("scenes"):
        lines.append("| *(none detected)* | | | | | |")
    lines.append("")

    lines.append("### Props")
    lines.append("")
    lines.append("| temporary_name | appears_in_blocks | evidence | confidence | confirmed_name | status |")
    lines.append("|---|---|---|---|---|---|")
    for entry in tables.get("props", []):
        lines.append(
            f"| {entry['temporary_name']} "
            f"| {', '.join(entry['appears_in_blocks'])} "
            f"| {entry['evidence']} "
            f"| {entry['confidence']} "
            f"| {entry['confirmed_name']} "
            f"| {entry['status']} |"
        )
    if not tables.get("props"):
        lines.append("| *(none detected)* | | | | | |")
    lines.append("")

    return "\n".join(lines)


def cmd_derive_naming_tables(args: argparse.Namespace) -> int:
    """Scan a filled draft_fill.json for temp: markers and generate naming_tables.json
    + optionally update cue_sheet.md with derived naming confirmation tables."""
    source_path = Path(args.source_json)
    if not source_path.exists():
        raise FileNotFoundError(f"Source JSON not found: {source_path}")

    source = json.loads(source_path.read_text(encoding="utf-8"))
    rows = source.get("rows", [])
    template = source.get("template", "production")

    if not rows:
        print("WARNING: No rows found in source JSON.", file=sys.stderr)

    tables = derive_naming_tables_from_rows(rows, template)

    # Count total markers found
    total = sum(len(entries) for entries in tables.values())

    # Output naming_tables.json
    output_path = Path(args.output)
    output_data = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source": str(source_path),
        "template": template,
        "total_temp_markers": total,
        "tables": tables,
    }
    write_json(output_path, output_data)
    print(f"Naming tables: {output_path} ({total} temp marker(s) found)")

    # Optionally update cue_sheet.md
    if args.md:
        md_path = Path(args.md)
        if not md_path.exists():
            print(f"WARNING: Markdown file not found: {md_path}", file=sys.stderr)
        else:
            md_content = md_path.read_text(encoding="utf-8")
            new_tables_md = format_naming_tables_md(tables)

            # Replace existing naming confirmation tables section if present
            # Look for "## Naming Confirmation Tables" ... next "## " section
            import re
            pattern = r"## Naming Confirmation Tables\n.*?(?=\n## |\Z)"
            if re.search(pattern, md_content, re.DOTALL):
                md_content = re.sub(pattern, new_tables_md.rstrip(), md_content, count=1, flags=re.DOTALL)
            else:
                # Append before "## Pending Questions" or at end
                pending_match = re.search(r"\n## Pending Questions", md_content)
                if pending_match:
                    md_content = (
                        md_content[:pending_match.start()]
                        + "\n" + new_tables_md + "\n"
                        + md_content[pending_match.start():]
                    )
                else:
                    md_content = md_content.rstrip() + "\n\n" + new_tables_md

            md_path.write_text(md_content, encoding="utf-8")
            print(f"Updated: {md_path}")

    # Summary
    for cat in ("characters", "scenes", "props"):
        count = len(tables.get(cat, []))
        if count:
            items = ", ".join(e["temporary_name"] for e in tables[cat])
            print(f"  {cat}: {count} ({items})")

    return 0


# ---------------------------------------------------------------------------
# LLM output normalize / lint
# ---------------------------------------------------------------------------

SHOT_SIZE_ENUM = {"WS", "MS", "CU", "EWS", "ECU"}
SHOT_SIZE_ALIASES: dict[str, str] = {
    "wide shot": "WS", "wide": "WS",
    "medium shot": "MS", "medium": "MS", "mid": "MS", "mid shot": "MS",
    "close-up": "CU", "close up": "CU", "closeup": "CU", "cu": "CU",
    "extreme wide shot": "EWS", "extreme wide": "EWS",
    "extreme close-up": "ECU", "extreme close up": "ECU", "extreme closeup": "ECU",
    "ws": "WS", "ms": "MS", "ews": "EWS", "ecu": "ECU",
}

MOTION_ENUM = {"static", "push-in", "pull-out", "pan", "tracking", "handheld"}
MOTION_ALIASES: dict[str, str] = {
    "push in": "push-in", "pushin": "push-in",
    "pull out": "pull-out", "pullout": "pull-out",
    "track": "tracking", "dolly": "tracking",
    "hand-held": "handheld", "hand held": "handheld",
    "still": "static", "locked": "static", "fixed": "static",
    "panning": "pan",
}

_VISUAL_HINT_RE = _re.compile(r"\[visual:\s*[^\]]*\]\s*")
_OCR_HINT_RE = _re.compile(r"\[OCR detected:\s*[^\]]*\]\s*")


def normalize_shot_size(value: str) -> tuple[str, bool]:
    """Normalize shot_size to enum value. Returns (normalized, was_changed)."""
    stripped = value.strip()
    if not stripped:
        return stripped, False
    upper = stripped.upper()
    if upper in SHOT_SIZE_ENUM:
        if stripped != upper:
            return upper, True
        return stripped, False
    lower = stripped.lower()
    if lower in SHOT_SIZE_ALIASES:
        return SHOT_SIZE_ALIASES[lower], True
    # Try matching with extra text like "WS (wide shot)"
    for enum_val in SHOT_SIZE_ENUM:
        if upper.startswith(enum_val):
            return enum_val, True
    return stripped, False


def normalize_motion(value: str) -> tuple[str, bool]:
    """Normalize motion to enum value. Returns (normalized, was_changed)."""
    stripped = value.strip()
    if not stripped:
        return stripped, False
    lower = stripped.lower()
    if lower in MOTION_ENUM:
        if stripped != lower:
            return lower, True
        return stripped, False
    if lower in MOTION_ALIASES:
        return MOTION_ALIASES[lower], True
    # Partial match: "slow push-in" -> "push-in"
    for enum_val in MOTION_ENUM:
        if enum_val in lower:
            return enum_val, True
    return stripped, False


def strip_hint_prefixes(value: str) -> tuple[str, bool]:
    """Remove [visual: ...] and [OCR detected: ...] hint prefixes."""
    if not value:
        return value, False
    cleaned = _VISUAL_HINT_RE.sub("", value)
    cleaned = _OCR_HINT_RE.sub("", cleaned)
    cleaned = cleaned.strip()
    return cleaned, cleaned != value.strip()


def cmd_normalize_fill(args: argparse.Namespace) -> int:
    """Normalize / lint a filled draft_fill.json or final_cues.json.

    Modes:
    - lint (default): report issues without modifying the file
    - fix: auto-normalize + report, write output
    """
    source_path = Path(args.source_json)
    if not source_path.exists():
        raise FileNotFoundError(f"Source JSON not found: {source_path}")

    source = json.loads(source_path.read_text(encoding="utf-8"))
    rows = source.get("rows", [])
    template = source.get("template", "production")
    fix_mode = bool(args.fix)

    issues: list[dict[str, Any]] = []
    fixes_applied: list[dict[str, Any]] = []

    columns = TEMPLATE_COLUMNS.get(template, TEMPLATE_COLUMNS["production"])

    for row in rows:
        block_id = row.get("shot_block", "?")

        # 1. Normalize shot_size
        if "shot_size" in row and row["shot_size"]:
            normalized, changed = normalize_shot_size(row["shot_size"])
            if changed:
                fix_rec = {
                    "block": block_id, "field": "shot_size",
                    "old": row["shot_size"], "new": normalized,
                    "type": "normalize",
                }
                if fix_mode:
                    row["shot_size"] = normalized
                    fixes_applied.append(fix_rec)
                else:
                    issues.append({**fix_rec, "severity": "fixable"})
            elif row["shot_size"].strip() and row["shot_size"].upper() not in SHOT_SIZE_ENUM:
                issues.append({
                    "block": block_id, "field": "shot_size",
                    "value": row["shot_size"],
                    "type": "unknown_enum",
                    "severity": "warning",
                    "message": f"shot_size '{row['shot_size']}' not in enum: {', '.join(sorted(SHOT_SIZE_ENUM))}",
                })

        # 2. Normalize motion
        if "motion" in row and row["motion"]:
            normalized, changed = normalize_motion(row["motion"])
            if changed:
                fix_rec = {
                    "block": block_id, "field": "motion",
                    "old": row["motion"], "new": normalized,
                    "type": "normalize",
                }
                if fix_mode:
                    row["motion"] = normalized
                    fixes_applied.append(fix_rec)
                else:
                    issues.append({**fix_rec, "severity": "fixable"})

        # 3. Strip [visual: ...] and [OCR detected: ...] hint prefixes
        for field in columns:
            value = row.get(field, "")
            if not isinstance(value, str):
                continue
            cleaned, changed = strip_hint_prefixes(value)
            if changed:
                fix_rec = {
                    "block": block_id, "field": field,
                    "old": value[:80], "new": cleaned[:80],
                    "type": "strip_hint",
                }
                if fix_mode:
                    row[field] = cleaned
                    fixes_applied.append(fix_rec)
                else:
                    issues.append({**fix_rec, "severity": "fixable"})

        # 4. Check temp: markers vs needs_confirmation
        needs_conf = str(row.get("needs_confirmation", "")).lower()
        naming_fields = ("scene", "characters", "location")
        for nf in naming_fields:
            value = str(row.get(nf, ""))
            markers = extract_temp_markers(value)
            for marker in markers:
                # Check if the marker (or a recognizable fragment) appears in needs_confirmation
                marker_key = marker.replace("temp:", "").strip().lower()
                if marker_key and marker_key not in needs_conf and "temp:" not in needs_conf:
                    issues.append({
                        "block": block_id, "field": nf,
                        "type": "orphaned_temp_marker",
                        "severity": "warning",
                        "message": f"'{marker}' in '{nf}' has no matching entry in needs_confirmation",
                    })

        # 5. Report empty required fields
        RECOMMENDED_FIELDS: dict[str, list[str]] = {
            "production": ["scene", "event", "shot_size", "mood", "characters"],
            "music-director": ["mood", "event", "music_note", "rhythm_change", "dynamics"],
            "script": ["scene", "event", "characters", "location"],
        }
        for rec_field in RECOMMENDED_FIELDS.get(template, []):
            if rec_field in set(columns) and not str(row.get(rec_field, "")).strip():
                issues.append({
                    "block": block_id, "field": rec_field,
                    "type": "empty_required",
                    "severity": "warning",
                    "message": f"Recommended field '{rec_field}' is empty",
                })

    # Build report
    report = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source": str(source_path),
        "template": template,
        "mode": "fix" if fix_mode else "lint",
        "total_rows": len(rows),
        "issues": issues,
        "fixes_applied": fixes_applied,
        "summary": {
            "fixable": len([i for i in issues if i.get("severity") == "fixable"]),
            "warnings": len([i for i in issues if i.get("severity") == "warning"]),
            "fixes_applied": len(fixes_applied),
        },
    }

    # Write output
    if fix_mode:
        out_path = Path(args.output) if args.output else source_path
        write_json(out_path, source)
        print(f"Fixed: {out_path} ({len(fixes_applied)} fix(es) applied)")

    if args.report_out:
        write_json(Path(args.report_out), report)

    # Print summary
    mode_label = "fix" if fix_mode else "lint"
    print(f"=== normalize-fill ({mode_label}, {template}) ===")
    print(f"Rows: {len(rows)}")
    if fixes_applied:
        print(f"Fixes applied: {len(fixes_applied)}")
        for f in fixes_applied[:10]:
            print(f"  {f['block']}.{f['field']}: '{f['old']}' -> '{f['new']}'")
        if len(fixes_applied) > 10:
            print(f"  ... and {len(fixes_applied) - 10} more")
    if issues:
        fixable = [i for i in issues if i.get("severity") == "fixable"]
        warnings = [i for i in issues if i.get("severity") == "warning"]
        if fixable:
            print(f"Fixable issues: {len(fixable)} (run with --fix to auto-normalize)")
            for i in fixable[:5]:
                print(f"  {i['block']}.{i['field']}: '{i.get('old', '')}' -> '{i.get('new', '')}'")
        if warnings:
            print(f"Warnings: {len(warnings)}")
            for i in warnings[:10]:
                print(f"  {i['block']}.{i['field']}: {i.get('message', i.get('type', ''))}")
    if not issues and not fixes_applied:
        print("  No issues found.")

    return 0


def cmd_merge_blocks(args: argparse.Namespace) -> int:
    analysis_path = Path(args.analysis_json)
    if not analysis_path.exists():
        raise FileNotFoundError(f"analysis.json not found: {analysis_path}")

    merge_plan_path = Path(args.merge_plan)
    if not merge_plan_path.exists():
        raise FileNotFoundError(f"Merge plan not found: {merge_plan_path}")

    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    merge_plan = json.loads(merge_plan_path.read_text(encoding="utf-8"))

    draft_blocks = analysis.get("draft_blocks", [])
    block_lookup: dict[str, dict[str, Any]] = {b["shot_block"]: b for b in draft_blocks}
    all_source_ids: set[str] = set(block_lookup.keys())

    merges = merge_plan.get("merges", [])
    errors: list[str] = []
    warnings: list[str] = []

    # --- Validate merge plan ---
    referenced_ids: list[str] = []
    seen_new_ids: set[str] = set()

    for idx, group in enumerate(merges):
        source_ids = group.get("source_blocks", [])
        new_id = group.get("new_id", f"A{idx+1}")

        if new_id in seen_new_ids:
            errors.append(f"Merge group {idx+1}: duplicate new_id '{new_id}'.")
        seen_new_ids.add(new_id)

        if not source_ids:
            errors.append(f"Merge group {idx+1} ('{new_id}'): source_blocks is empty.")

        for sid in source_ids:
            if sid not in block_lookup:
                errors.append(f"Merge group {idx+1} ('{new_id}'): source block '{sid}' not found in analysis.")
            if sid in referenced_ids:
                errors.append(f"Merge group {idx+1} ('{new_id}'): source block '{sid}' already used in another group.")
            referenced_ids.append(sid)

    unreferenced = all_source_ids - set(referenced_ids)
    strict = bool(args.strict) if hasattr(args, "strict") else False
    if unreferenced:
        if strict:
            errors.append(f"Blocks not referenced in merge plan: {', '.join(sorted(unreferenced))}. In strict mode this is an error.")
        else:
            warnings.append(f"Blocks not referenced in merge plan: {', '.join(sorted(unreferenced))}. They will be auto-appended as unmerged blocks.")

    # --- If errors, report and stop ---
    if errors:
        report = {
            "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
            "valid": False,
            "errors": errors,
            "warnings": warnings,
        }
        if args.output:
            write_json(Path(args.output), report)
        print("=== merge-blocks validation FAILED ===", file=sys.stderr)
        for e in errors:
            print(f"  ✗ {e}", file=sys.stderr)
        for w in warnings:
            print(f"  ⚠ {w}", file=sys.stderr)
        return 1

    # --- Perform merge ---
    merged_blocks: list[dict[str, Any]] = []
    for idx, group in enumerate(merges, start=1):
        source_ids = group.get("source_blocks", [])
        new_id = group.get("new_id", f"A{idx}")
        sources = [block_lookup[sid] for sid in source_ids if sid in block_lookup]
        if not sources:
            continue

        start_seconds = min(s["start_seconds"] for s in sources)
        end_seconds = max(s["end_seconds"] for s in sources)
        keyframe = group.get("keyframe") or sources[0].get("keyframe")

        merged_blocks.append({
            "shot_block": new_id,
            "start_seconds": start_seconds,
            "start_time": format_seconds(start_seconds),
            "end_seconds": end_seconds,
            "end_time": format_seconds(end_seconds),
            "keyframe": keyframe,
            "source_blocks": source_ids,
            "merge_reason": group.get("reason", ""),
        })

    # --- Auto-append unreferenced blocks (non-strict mode) ---
    if unreferenced and not strict:
        unref_blocks = sorted(
            [block_lookup[sid] for sid in unreferenced if sid in block_lookup],
            key=lambda b: b["start_seconds"],
        )
        for block in unref_blocks:
            merged_blocks.append({
                "shot_block": block["shot_block"],
                "start_seconds": block["start_seconds"],
                "start_time": block["start_time"],
                "end_seconds": block["end_seconds"],
                "end_time": block["end_time"],
                "keyframe": block.get("keyframe"),
                "source_blocks": [block["shot_block"]],
                "merge_reason": "auto-appended: not referenced in merge plan",
                "unmerged": True,
            })

    # --- Check output ordering ---
    for i in range(1, len(merged_blocks)):
        prev_end = merged_blocks[i - 1]["end_seconds"]
        curr_start = merged_blocks[i]["start_seconds"]
        if curr_start < prev_end - 0.05:
            msg = (
                f"Block '{merged_blocks[i]['shot_block']}' starts at {merged_blocks[i]['start_time']} "
                f"which overlaps with previous block ending at {merged_blocks[i-1]['end_time']}."
            )
            if strict:
                errors.append(msg)
            else:
                warnings.append(msg)

    # --- If strict errors found in post-merge checks, report and fail ---
    if errors:
        report = {
            "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
            "valid": False,
            "strict": strict,
            "errors": errors,
            "warnings": warnings,
        }
        if args.output:
            write_json(Path(args.output), report)
        print("=== merge-blocks validation FAILED ===", file=sys.stderr)
        for e in errors:
            print(f"  ✗ {e}", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    output_data = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source_analysis": str(analysis_path),
        "merge_plan": str(merge_plan_path),
        "original_block_count": len(draft_blocks),
        "merged_block_count": len(merged_blocks),
        "unreferenced_blocks": sorted(unreferenced),
        "warnings": warnings,
        "blocks": merged_blocks,
    }
    write_json(output_path, output_data)

    if warnings:
        for w in warnings:
            print(f"  ⚠ {w}", file=sys.stderr)

    print(str(output_path))
    return 0


def cmd_export_md(args: argparse.Namespace) -> int:
    cue_json = Path(args.cue_json)
    if not cue_json.exists():
        raise FileNotFoundError(f"Cue JSON not found: {cue_json}")

    payload = json.loads(cue_json.read_text(encoding="utf-8"))
    template = args.template or payload.get("template") or "production"
    if template not in TEMPLATE_COLUMNS:
        raise ValueError(f"Unknown template: {template}")

    base_dir = Path(args.base_dir).resolve() if hasattr(args, "base_dir") and args.base_dir else cue_json.parent.resolve()
    rows = payload.get("rows", [])
    columns = TEMPLATE_COLUMNS[template]
    output_path = Path(args.output)
    ensure_parent(output_path)

    lines: list[str] = []
    title = payload.get("video_title", "Untitled")
    lines.append(f"# Cue Sheet — {title} ({template})")
    lines.append("")

    lines.append("## Video Info")
    lines.append("")
    lines.append(f"- **Source**: `{payload.get('source_path', '')}`")
    lines.append(f"- **Generated**: {payload.get('generated_at', '')}")
    lines.append(f"- **Template**: {template}")
    lines.append(f"- **Blocks**: {len(rows)}")
    lines.append("")

    lines.append("## Shot Blocks")
    lines.append("")

    header = "| " + " | ".join(columns) + " |"
    separator = "|" + "|".join(["---"] * len(columns)) + "|"
    lines.append(header)
    lines.append(separator)

    # --- Delivery completeness tracking ---
    missing_keyframes_md: list[str] = []

    for row in rows:
        cells = []
        for col in columns:
            value = str(row.get(col, ""))
            if col == "keyframe" and value:
                # Resolve through base_dir first, then compute relative to output
                kf_path = resolve_keyframe_path(base_dir, value)
                if kf_path and not kf_path.exists():
                    missing_keyframes_md.append(row.get("shot_block", "?"))
                try:
                    rel = os.path.relpath(str(kf_path or value), output_path.parent).replace("\\", "/")
                except Exception:
                    rel = value.replace("\\", "/")
                cells.append(f"![kf]({rel})")
            else:
                cells.append(value.replace("|", "\\|").replace("\n", " "))
        lines.append("| " + " | ".join(cells) + " |")

    lines.append("")

    has_unconfirmed = any(
        row.get("needs_confirmation")
        for row in rows
    )
    if has_unconfirmed:
        lines.append("## Pending Confirmation")
        lines.append("")
        for row in rows:
            nc = row.get("needs_confirmation", "")
            if nc:
                lines.append(f"- **{row.get('shot_block', '?')}**: {nc}")
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(str(output_path))

    # --- Delivery completeness summary ---
    if missing_keyframes_md:
        print(f"WARNING: keyframe files not found for blocks: {', '.join(missing_keyframes_md)}", file=sys.stderr)
    return 0


def cmd_build_final_skeleton(args: argparse.Namespace) -> int:
    """Generate a final_cues.json skeleton from merged blocks, analysis draft_blocks,
    or a filled draft_fill.json. Fields are populated with empty strings for LLM to
    fill in (unless source already has filled content, e.g. from draft_fill.json)."""
    source_path = Path(args.source_json)
    if not source_path.exists():
        raise FileNotFoundError(f"Source JSON not found: {source_path}")

    source = json.loads(source_path.read_text(encoding="utf-8"))

    # Accept multiple input formats:
    # 1. draft_fill.json (has "rows" with template columns + "fill_status")
    # 2. merged_blocks.json (has "blocks")
    # 3. analysis.json (has "draft_blocks")
    is_fill_input = "fill_status" in source
    if is_fill_input and source.get("rows"):
        # draft_fill.json — rows already have template column structure
        blocks = source["rows"]
        fill_status = source.get("fill_status", "unknown")
        if fill_status == "partial":
            # Check how many content fields are still empty, using the actual
            # template columns (not a hardcoded subset) to catch all templates
            STRUCTURAL_FIELDS = {"shot_block", "start_time", "end_time", "keyframe"}
            active_template = source.get("template", "production")
            template_cols = set(TEMPLATE_COLUMNS.get(active_template, TEMPLATE_COLUMNS["production"]))
            content_cols = template_cols - STRUCTURAL_FIELDS
            empty_count = 0
            for row in blocks:
                for f in content_cols:
                    val = row.get(f, "")
                    if not val or val.startswith("[visual:") or val.startswith("[OCR detected:"):
                        empty_count += 1
            if empty_count > 0:
                print(f"WARNING: fill_status is 'partial' — {empty_count} content field(s) still empty or hint-only "
                      f"(template={active_template}, {len(content_cols)} content fields per block). "
                      f"LLM fill-in may be incomplete.", file=sys.stderr)
        elif fill_status == "pending":
            print("WARNING: fill_status is 'pending' — no LLM fill-in has been done. "
                  "The final JSON will have mostly empty content fields.", file=sys.stderr)
    else:
        blocks = source.get("blocks") or source.get("draft_blocks", [])

    template = args.template or source.get("template") or "production"
    if template not in TEMPLATE_COLUMNS:
        raise ValueError(f"Unknown template: {template}")

    columns = TEMPLATE_COLUMNS[template]
    rows: list[dict[str, Any]] = []

    for block in blocks:
        row: dict[str, Any] = {}
        for col in columns:
            if col == "shot_block":
                row[col] = block.get("shot_block", "")
            elif col == "start_time":
                row[col] = block.get("start_time", "")
            elif col == "end_time":
                row[col] = block.get("end_time", "")
            elif col == "keyframe":
                row[col] = block.get("keyframe", "")
            else:
                # Preserve existing content if source already has it (draft_fill.json)
                existing = block.get(col, "")
                row[col] = existing if existing else ""
        rows.append(row)

    # Resolve metadata with fallback + warning
    # Priority: CLI arg > source video_title > source video.source_path (stem only)
    raw_title = args.video_title or source.get("video_title") or source.get("video", {}).get("source_path")
    if raw_title and not args.video_title and not source.get("video_title"):
        # Came from source_path — normalize to stem
        video_title = Path(raw_title).stem
    elif raw_title:
        video_title = raw_title
    else:
        print("WARNING: --video-title not provided and source JSON has no video metadata. "
              "Falling back to 'untitled'. Consider passing --video-title explicitly.", file=sys.stderr)
        video_title = "untitled"
    source_path_value = args.source_path or source.get("source_path") or source.get("video", {}).get("source_path") or source.get("source_analysis", "")
    if not source_path_value:
        print("WARNING: --source-path not provided and source JSON has no video metadata. "
              "Falling back to source_analysis path. Consider passing --source-path explicitly.", file=sys.stderr)

    skeleton = {
        "template": template,
        "video_title": video_title,
        "source_path": source_path_value,
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
    }

    output_path = Path(args.output)
    write_json(output_path, skeleton)
    print(str(output_path))
    return 0


def compute_block_continuity(
    block_a: dict[str, Any],
    block_b: dict[str, Any],
    sampled_frames: list[dict[str, Any]],
    asr_segments: list[dict[str, Any]],
    threshold: float = 0.65,
) -> dict[str, Any]:
    """Compute a continuity score between two adjacent blocks.
    Higher score = more likely candidates for merging."""
    scores: dict[str, float] = {}

    # 1. Visual similarity: histogram distance between keyframes via sampled_frames
    vf_a = block_a.get("visual_features") or {}
    vf_b = block_b.get("visual_features") or {}
    if vf_a and vf_b:
        brightness_diff = abs(vf_a.get("brightness", 128) - vf_b.get("brightness", 128)) / 255.0
        contrast_diff = abs(vf_a.get("contrast", 50) - vf_b.get("contrast", 50)) / 128.0
        saturation_diff = abs(vf_a.get("saturation", 50) - vf_b.get("saturation", 50)) / 255.0
        hue_diff = abs(vf_a.get("dominant_hue", 90) - vf_b.get("dominant_hue", 90)) / 180.0
        visual_sim = 1.0 - (brightness_diff * 0.3 + contrast_diff * 0.2 + saturation_diff * 0.2 + hue_diff * 0.3)
        scores["visual_similarity"] = round(max(visual_sim, 0.0), 3)
        # Same tone = bonus
        scores["same_tone"] = 1.0 if vf_a.get("tone") == vf_b.get("tone") else 0.0
        scores["same_color_temp"] = 1.0 if vf_a.get("color_temp") == vf_b.get("color_temp") else 0.0
    else:
        scores["visual_similarity"] = 0.5  # neutral if no data

    # 2. Cut strength: low histogram distance at boundary suggests visual continuity
    b_score = block_b.get("candidate_score")
    if isinstance(b_score, (int, float)):
        scores["cut_weakness"] = round(1.0 - min(b_score, 1.0), 3)
    else:
        scores["cut_weakness"] = 0.5

    # 3. ASR continuity: dialogue spanning the boundary
    boundary = block_a.get("end_seconds", 0.0)
    tolerance = 1.0  # seconds
    asr_spans = any(
        s.get("start", 0) < boundary + tolerance and s.get("end", 0) > boundary - tolerance
        for s in asr_segments
    )
    scores["asr_continuity"] = 1.0 if asr_spans else 0.0

    # 4. Temporal proximity: very short blocks are merge candidates
    dur_a = block_a.get("end_seconds", 0) - block_a.get("start_seconds", 0)
    dur_b = block_b.get("end_seconds", 0) - block_b.get("start_seconds", 0)
    short_block = dur_a < 1.5 or dur_b < 1.5
    scores["short_block_bonus"] = 0.3 if short_block else 0.0

    # Weighted total
    total = (
        scores.get("visual_similarity", 0.5) * 0.35
        + scores.get("cut_weakness", 0.5) * 0.25
        + scores.get("same_tone", 0) * 0.1
        + scores.get("same_color_temp", 0) * 0.05
        + scores.get("asr_continuity", 0) * 0.15
        + scores.get("short_block_bonus", 0) * 0.1
    )

    return {
        "block_a": block_a.get("shot_block", ""),
        "block_b": block_b.get("shot_block", ""),
        "continuity_score": round(total, 3),
        "component_scores": scores,
        "suggest_merge": total >= threshold,
    }


def cmd_suggest_merges(args: argparse.Namespace) -> int:
    """Compute inter-block continuity scores and suggest merge candidates.
    Outputs a preliminary merge plan that the LLM can review and adjust."""
    analysis_path = Path(args.analysis_json)
    if not analysis_path.exists():
        raise FileNotFoundError(f"analysis.json not found: {analysis_path}")

    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    draft_blocks = analysis.get("draft_blocks", [])
    sampled_frames = analysis.get("sampled_frames", [])
    asr_segments = analysis.get("asr", {}).get("segments", [])

    if len(draft_blocks) < 2:
        print("Not enough blocks to suggest merges (need at least 2).", file=sys.stderr)
        return 0

    threshold = float(args.threshold) if hasattr(args, "threshold") and args.threshold else 0.65

    # Compute pairwise continuity
    pairs: list[dict[str, Any]] = []
    for i in range(len(draft_blocks) - 1):
        pair = compute_block_continuity(
            draft_blocks[i], draft_blocks[i + 1], sampled_frames, asr_segments,
            threshold=threshold,
        )
        pairs.append(pair)

    # Build suggested merge groups from consecutive high-continuity pairs
    merge_groups: list[dict[str, Any]] = []
    current_group: list[str] = [draft_blocks[0]["shot_block"]]
    current_reasons: list[str] = []

    for pair in pairs:
        if pair["suggest_merge"]:
            current_group.append(pair["block_b"])
            scores = pair["component_scores"]
            reasons = []
            if scores.get("visual_similarity", 0) > 0.7:
                reasons.append("visually similar")
            if scores.get("asr_continuity", 0) > 0:
                reasons.append("dialogue spans boundary")
            if scores.get("short_block_bonus", 0) > 0:
                reasons.append("short block")
            if scores.get("cut_weakness", 0) > 0.6:
                reasons.append("weak cut boundary")
            current_reasons.extend(reasons)
        else:
            # Flush current group
            if len(current_group) > 1:
                merge_groups.append({
                    "source_blocks": current_group,
                    "new_id": current_group[0],
                    "keyframe": None,
                    "reason": f"auto-suggested: {'; '.join(set(current_reasons)) if current_reasons else 'high continuity score'}",
                    "confidence": "auto",
                })
            current_group = [pair["block_b"]]
            current_reasons = []

    # Flush last group
    if len(current_group) > 1:
        merge_groups.append({
            "source_blocks": current_group,
            "new_id": current_group[0],
            "keyframe": None,
            "reason": f"auto-suggested: {'; '.join(set(current_reasons)) if current_reasons else 'high continuity score'}",
            "confidence": "auto",
        })

    # Also include singleton blocks that weren't merged
    merged_ids = set()
    for g in merge_groups:
        merged_ids.update(g["source_blocks"])
    singletons = [b["shot_block"] for b in draft_blocks if b["shot_block"] not in merged_ids]
    for sid in singletons:
        merge_groups.append({
            "source_blocks": [sid],
            "new_id": sid,
            "keyframe": None,
            "reason": "no merge suggested — keep as separate block",
            "confidence": "auto",
        })

    # Renumber
    for idx, group in enumerate(merge_groups, start=1):
        group["new_id"] = make_block_id(idx)

    output = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "_instructions": (
            "This is an AUTO-GENERATED merge suggestion based on visual continuity scoring. "
            "The LLM should REVIEW and ADJUST this plan before passing to merge-blocks. "
            "Specifically: check that narrative function boundaries are respected "
            "(scene changes, flashback transitions, format changes should NOT be merged)."
        ),
        "threshold": threshold,
        "total_blocks": len(draft_blocks),
        "suggested_merge_groups": len([g for g in merge_groups if len(g["source_blocks"]) > 1]),
        "pairwise_scores": pairs,
        "merges": merge_groups,
    }

    output_path = Path(args.output)
    write_json(output_path, output)

    # Print summary
    merged_count = sum(1 for g in merge_groups if len(g["source_blocks"]) > 1)
    kept_count = sum(1 for g in merge_groups if len(g["source_blocks"]) == 1)
    print(f"Suggested merge plan: {merged_count} merge group(s), {kept_count} kept separate.")
    print(f"Output: {output_path}")
    if merged_count > 0:
        print("Merge suggestions:")
        for g in merge_groups:
            if len(g["source_blocks"]) > 1:
                print(f"  {' + '.join(g['source_blocks'])} → {g['new_id']} ({g['reason']})")
    print("NOTE: LLM should review this plan for narrative-function boundaries before executing.")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="cuesheet-creator minimal toolkit")
    parser.add_argument("--version", action="version", version=f"cuesheet-creator {__version__}")
    parser.add_argument("--ffmpeg-path", type=resolved_path, default=None, help="Explicit path to ffmpeg executable (overrides auto-detection)")
    parser.add_argument("--ffprobe-path", type=resolved_path, default=None, help="Explicit path to ffprobe executable (overrides auto-detection)")
    subparsers = parser.add_subparsers(dest="command", required=True)

    selfcheck = subparsers.add_parser("selfcheck", help="Check runtime environment")
    selfcheck.add_argument("--json-out", type=resolved_path, help="Write JSON report to file")
    selfcheck.add_argument("--output-format", choices=["text", "json"], default="text")
    selfcheck.add_argument("--fail-on-missing-required", action="store_true")
    selfcheck.set_defaults(func=cmd_selfcheck)

    install = subparsers.add_parser("install-deps", help="Install missing Python dependencies")
    install.add_argument(
        "--include-optional",
        default="none",
        help="Optional dependency groups: none / scene / asr / ocr / all (comma-separated)",
    )
    install.add_argument("--dry-run", action="store_true", help="Output install plan only, do not install")
    install.add_argument("--report-out", type=resolved_path, help="Write install report JSON")
    install.add_argument("--output-format", choices=["text", "json"], default="text")
    install.add_argument("--index-url", help="pip index-url (e.g. domestic mirror)")
    install.add_argument("--extra-index-url", help="pip extra-index-url")
    install.add_argument("--upgrade-pip", action="store_true", help="Run pip install --upgrade pip first")
    install.add_argument("--fail-on-blocking", action="store_true", help="Return non-zero if blocking issues remain after install")
    install.set_defaults(func=cmd_install_deps)

    prepare = subparsers.add_parser("prepare-env", help="One-command env check, install, and recheck")
    prepare.add_argument(
        "--mode",
        choices=sorted(PREPARE_ENV_MODES),
        default="check-only",
        help="Prepare mode: check-only / install-required / install-scene / install-asr / install-ocr / install-all",
    )
    prepare.add_argument("--out-dir", type=resolved_path, help="Output directory; auto-writes prepare_env.json / selfcheck.pre.json / selfcheck.post.json")
    prepare.add_argument("--dry-run", action="store_true", help="Output plan only, do not install")
    prepare.add_argument("--selfcheck-out", type=resolved_path, help="Write pre-check JSON (overrides --out-dir default)")

    prepare.add_argument("--install-report-out", type=resolved_path, help="Write install-deps report JSON (overrides --out-dir default)")
    prepare.add_argument("--postcheck-out", type=resolved_path, help="Write post-check JSON (overrides --out-dir default)")
    prepare.add_argument("--report-out", type=resolved_path, help="Write prepare-env summary JSON (overrides --out-dir default)")

    prepare.add_argument("--output-format", choices=["text", "json"], default="text")
    prepare.add_argument("--index-url", help="pip index-url (e.g. domestic mirror)")
    prepare.add_argument("--extra-index-url", help="pip extra-index-url")
    prepare.add_argument("--upgrade-pip", action="store_true", help="Run pip install --upgrade pip first")
    prepare.add_argument("--fail-on-blocking", action="store_true", help="Return non-zero if blocking issues remain")
    prepare.set_defaults(func=cmd_prepare_env)

    scan = subparsers.add_parser("scan-video", help="Extract frames and generate analysis.json")

    scan.add_argument("--video", type=resolved_path, required=True, help="Video file path")
    scan.add_argument("--out-dir", type=resolved_path, default=None, help="Output directory (default: <video-dir>/<video-name>_cuesheet/)")
    scan.add_argument("--sample-interval", type=float, default=2.0, help="Sampling interval in seconds")
    scan.add_argument("--scene-threshold", type=float, default=0.35, help="Histogram cut threshold (fallback mode)")
    scan.add_argument("--content-threshold", type=float, default=27.0, help="PySceneDetect ContentDetector threshold")
    scan.add_argument("--max-samples", type=int, default=0, help="Max sample count, 0 for unlimited")
    scan.add_argument("--max-width", type=int, default=1280, help="Max keyframe export width")
    scan.add_argument("--asr", action="store_true", help="Enable ASR speech recognition (requires faster-whisper)")
    scan.add_argument("--asr-model", default="base", help="ASR model size: tiny / base / small / medium / large-v3")
    scan.add_argument("--ocr", action="store_true", help="Enable OCR text detection (requires rapidocr / easyocr / paddleocr)")
    scan.add_argument("--start-time", default=None, help="Clip start time (HH:MM:SS.mmm, HH:MM:SS, MM:SS, or seconds)")
    scan.add_argument("--end-time", default=None, help="Clip end time (HH:MM:SS.mmm, HH:MM:SS, MM:SS, or seconds)")
    scan.add_argument("--output-format", choices=["text", "json"], default="text", help="Output format: text (human) or json (agent-friendly summary)")
    scan.set_defaults(func=cmd_scan_video)

    draft = subparsers.add_parser("draft-from-analysis", help="Generate draft skeleton from analysis.json")
    draft.add_argument("--analysis-json", type=resolved_path, required=True, help="Path to analysis.json")
    draft.add_argument("--output", type=resolved_path, required=True, help="Output Markdown path")
    draft.add_argument(
        "--template",
        choices=["production", "music-director", "script"],
        default="production",
        help="Draft template type",
    )
    draft.set_defaults(func=cmd_draft_from_analysis)

    build = subparsers.add_parser("build-xlsx", help="Export Excel from final_cues.json")
    build.add_argument("--cue-json", type=resolved_path, required=True, help="Structured cue JSON")
    build.add_argument("--output", type=resolved_path, required=True, help="Output xlsx path")
    build.add_argument("--base-dir", type=resolved_path, help="Base directory for resolving relative keyframe paths")
    build.add_argument(
        "--template",
        choices=["production", "music-director", "script"],
        help="Override template in cue JSON",
    )
    build.add_argument("--image-max-width", type=int, default=180)
    build.add_argument("--image-max-height", type=int, default=100)
    build.set_defaults(func=cmd_build_xlsx)

    validate = subparsers.add_parser("validate-cue-json", help="Validate final_cues.json structural integrity")
    validate.add_argument("--cue-json", type=resolved_path, required=True, help="Cue JSON to validate")
    validate.add_argument(
        "--template",
        choices=["production", "music-director", "script"],
        help="Override template for field validation",
    )
    validate.add_argument("--report-out", type=resolved_path, help="Write validation report JSON")
    validate.add_argument("--base-dir", type=resolved_path, help="Base directory for resolving relative keyframe paths (used with --check-files)")
    validate.add_argument("--check-files", action="store_true", help="Verify that keyframe files actually exist on disk")
    validate.add_argument("--output-format", choices=["text", "json"], default="text")
    validate.set_defaults(func=cmd_validate_cue_json)

    apply_naming = subparsers.add_parser("apply-naming", help="Batch-apply naming overrides")
    apply_naming.add_argument("--overrides", type=resolved_path, required=True, help="Naming overrides JSON file")
    apply_naming.add_argument("--cue-json", type=resolved_path, help="final_cues.json to apply replacements to")
    apply_naming.add_argument("--md", type=resolved_path, help="cue_sheet.md to apply replacements to")
    apply_naming.add_argument("--output", type=resolved_path, help="Write updated JSON to this path instead of overwriting original")
    apply_naming.add_argument("--dry-run", action="store_true", help="Preview changes without writing files")
    apply_naming.add_argument("--report-out", type=resolved_path, help="Write replacement report JSON")
    apply_naming.set_defaults(func=cmd_apply_naming)

    derive_naming = subparsers.add_parser("derive-naming-tables", help="Scan filled draft_fill.json for temp: markers and generate naming confirmation tables")
    derive_naming.add_argument("--source-json", type=resolved_path, required=True, help="Filled draft_fill.json (or any JSON with rows containing temp: markers)")
    derive_naming.add_argument("--output", type=resolved_path, required=True, help="Output naming_tables.json path")
    derive_naming.add_argument("--md", type=resolved_path, help="cue_sheet.md to update with derived naming tables (optional)")
    derive_naming.set_defaults(func=cmd_derive_naming_tables)

    normalize = subparsers.add_parser("normalize-fill", help="Normalize/lint LLM-filled JSON: standardize enums, strip hint prefixes, check temp markers")
    normalize.add_argument("--source-json", type=resolved_path, required=True, help="Filled draft_fill.json or final_cues.json")
    normalize.add_argument("--fix", action="store_true", help="Auto-normalize and write output (default: lint-only, report issues)")
    normalize.add_argument("--output", type=resolved_path, help="Output path for fixed JSON (default: overwrite source)")
    normalize.add_argument("--report-out", type=resolved_path, help="Write normalize report JSON")
    normalize.set_defaults(func=cmd_normalize_fill)

    merge = subparsers.add_parser("merge-blocks", help="Merge draft blocks based on a merge plan")
    merge.add_argument("--analysis-json", type=resolved_path, required=True, help="Path to analysis.json")
    merge.add_argument("--merge-plan", type=resolved_path, required=True, help="Path to merge plan JSON")
    merge.add_argument("--output", type=resolved_path, required=True, help="Output merged blocks JSON")
    merge.add_argument("--strict", action="store_true", help="Fail on unreferenced blocks and time ordering issues")
    merge.set_defaults(func=cmd_merge_blocks)

    suggest = subparsers.add_parser("suggest-merges", help="Auto-suggest block merges based on visual continuity scoring")
    suggest.add_argument("--analysis-json", type=resolved_path, required=True, help="Path to analysis.json")
    suggest.add_argument("--output", type=resolved_path, required=True, help="Output suggested merge plan JSON")
    suggest.add_argument("--threshold", type=float, default=0.65, help="Continuity score threshold for merge suggestion (0.0-1.0, default 0.65)")
    suggest.set_defaults(func=cmd_suggest_merges)

    export_md = subparsers.add_parser("export-md", help="Generate Markdown final from final_cues.json")
    export_md.add_argument("--cue-json", type=resolved_path, required=True, help="Structured cue JSON")
    export_md.add_argument("--output", type=resolved_path, required=True, help="Output Markdown path")
    export_md.add_argument("--base-dir", type=resolved_path, help="Base directory for resolving relative keyframe paths")
    export_md.add_argument(
        "--template",
        choices=["production", "music-director", "script"],
        help="Override template in cue JSON",
    )
    export_md.set_defaults(func=cmd_export_md)

    skeleton = subparsers.add_parser("build-final-skeleton", help="Generate empty final_cues.json skeleton from merged/draft blocks")
    skeleton.add_argument("--source-json", type=resolved_path, required=True, help="Merged blocks JSON or analysis.json")
    skeleton.add_argument("--output", type=resolved_path, required=True, help="Output final_cues.json skeleton path")
    skeleton.add_argument(
        "--template",
        choices=["production", "music-director", "script"],
        default="production",
        help="Template for field selection",
    )
    skeleton.add_argument("--video-title", default=None, help="Video title for the skeleton")
    skeleton.add_argument("--source-path", type=resolved_path, default=None, help="Video source path override (for merged input that lacks video metadata)")
    skeleton.set_defaults(func=cmd_build_final_skeleton)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    # Apply CLI overrides for ffmpeg/ffprobe paths
    if args.ffmpeg_path:
        _CLI_COMMAND_OVERRIDES["ffmpeg"] = args.ffmpeg_path
    if args.ffprobe_path:
        _CLI_COMMAND_OVERRIDES["ffprobe"] = args.ffprobe_path
    try:
        return int(args.func(args))
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
