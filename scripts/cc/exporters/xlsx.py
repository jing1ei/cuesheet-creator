"""Excel export with embedded keyframes."""
from __future__ import annotations

import datetime as dt
import json
import tempfile
from pathlib import Path

from cc.constants import DEFAULT_COLUMN_WIDTHS, TEMPLATE_COLUMNS
from cc.templates import get_template_column_widths, validate_template_name
from cc.utils import ensure_parent, read_json, resolve_keyframe_path, scale_dimensions
from cc.validation import evaluate_delivery_readiness


def cmd_build_xlsx(args: "argparse.Namespace") -> int:  # noqa: F821
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from PIL import Image as PILImage
    except Exception as exc:
        raise RuntimeError(f"Modules required for Excel export unavailable: {exc}") from exc

    cue_json = Path(args.cue_json)
    if not cue_json.exists():
        raise FileNotFoundError(f"Cue JSON not found: {cue_json}")

    payload = read_json(cue_json)
    template = args.template or payload.get("template") or "production"
    validate_template_name(template)

    base_dir = Path(args.base_dir).resolve() if args.base_dir else cue_json.parent.resolve()
    rows = payload.get("rows", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Cue Sheet"
    meta = wb.create_sheet("Meta")

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    columns = TEMPLATE_COLUMNS[template]

    # --embed-keyframes: inject a keyframe column even if the template doesn't define one
    embed_keyframes_override = hasattr(args, "embed_keyframes") and args.embed_keyframes
    if embed_keyframes_override and "keyframe" not in columns:
        # Insert keyframe column right after end_time (position 3, 0-indexed)
        columns = list(columns)  # copy to avoid mutating the registry
        insert_pos = columns.index("end_time") + 1 if "end_time" in columns else 3
        columns.insert(insert_pos, "keyframe")

    tmpl_widths = get_template_column_widths(template)
    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment
        ws.column_dimensions[cell.column_letter].width = tmpl_widths.get(column_name, DEFAULT_COLUMN_WIDTHS.get(column_name, 18))

    keyframe_col_index = columns.index("keyframe") + 1 if "keyframe" in columns else None

    embedded_keyframes = 0
    missing_keyframes: list[str] = []
    thumb_dir = tempfile.mkdtemp(prefix="cuesheet_thumbs_") if keyframe_col_index is not None else None

    for row_idx, item in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(columns, start=1):
            value = item.get(column_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value if column_name != "keyframe" else "")
            cell.alignment = wrap_alignment

        if keyframe_col_index is not None:
            keyframe_value = item.get("keyframe") or item.get("_keyframe")
            keyframe_path = resolve_keyframe_path(base_dir, keyframe_value)
            if keyframe_path and keyframe_path.exists():
                with PILImage.open(keyframe_path) as img:
                    width, height = img.size
                    scaled_w, scaled_h = scale_dimensions(width, height, args.image_max_width, args.image_max_height)
                    _lanczos = getattr(PILImage, "Resampling", PILImage).LANCZOS
                    thumb = img.resize((scaled_w, scaled_h), _lanczos)
                    thumb_path = Path(thumb_dir) / f"thumb_{row_idx}_{keyframe_path.name}"
                    thumb.save(str(thumb_path), "JPEG", quality=85)
                xl_image = XLImage(str(thumb_path))
                xl_image.width = scaled_w
                xl_image.height = scaled_h
                anchor_cell = ws.cell(row=row_idx, column=keyframe_col_index)
                ws.add_image(xl_image, anchor_cell.coordinate)
                anchor_cell.value = keyframe_path.name
                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, scaled_h * 0.78)
                embedded_keyframes += 1
            else:
                block_label = item.get("shot_block", f"row{row_idx - 1}")
                missing_keyframes.append(block_label)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta_rows = [
        ("template", template),
        ("video_title", payload.get("video_title", "")),
        ("source_path", payload.get("source_path", "")),
        ("generated_at", payload.get("generated_at", dt.datetime.now().isoformat(timespec="seconds"))),
        ("row_count", len(rows)),
    ]
    for index, (key, value) in enumerate(meta_rows, start=1):
        meta.cell(row=index, column=1, value=key)
        meta.cell(row=index, column=2, value=value)

    output_path = Path(args.output)
    ensure_parent(output_path)
    try:
        wb.save(output_path)
    finally:
        # Clean up temp thumbnails regardless of save success/failure —
        # they live in an isolated temp directory, not next to keyframes
        if thumb_dir:
            import shutil
            shutil.rmtree(thumb_dir, ignore_errors=True)

    delivery = evaluate_delivery_readiness(
        rows, template, base_dir=base_dir, check_files=True,
    )
    summary = {
        "status": "ok",
        "stage": "build-xlsx",
        "output": str(output_path),
        "template": template,
        "embedded_keyframes": embedded_keyframes,
        **delivery,
    }
    if hasattr(args, "output_format") and args.output_format == "json":
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print(str(output_path))
        print("--- delivery summary ---")
        print(f"  rows exported: {delivery['row_count']}")
        if not rows:
            print("  WARNING: no rows exported -- cue sheet is empty")
        if keyframe_col_index is not None:
            print(f"  embedded keyframes: {embedded_keyframes}/{len(rows)}")
            if delivery["missing_keyframes"]:
                print(f"  WARNING: missing keyframes for blocks: {', '.join(delivery['missing_keyframes'])}")
        if delivery["empty_required_fields"] > 0:
            print(f"  WARNING: {delivery['empty_required_fields']} empty required field(s) across all rows")
        if delivery["empty_recommended_fields"] > 0:
            print(f"  WARNING: {delivery['empty_recommended_fields']} empty recommended field(s) across all rows")
        print(f"  delivery_ready: {'YES' if delivery['delivery_ready'] else 'NO'}")

    fail_on_gap = hasattr(args, "fail_on_delivery_gap") and args.fail_on_delivery_gap
    if fail_on_gap and not delivery["delivery_ready"]:
        return 1
    return 0


__all__ = ["cmd_build_xlsx"]
